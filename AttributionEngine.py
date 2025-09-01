import pandas as pd
import numpy as np
import json
import logging
from datetime import datetime, timedelta
import re
from typing import Dict, List, Optional, Tuple
import psycopg2
from sqlalchemy import create_engine, text

# Set up logging
logging.basicConfig(
    filename='attribution_analysis.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    filemode='w',
    force=True
)

class AttributionEngine:
    def __init__(self, time_range_days: int = 30, db_config: Dict = None, 
                 start_date: str = None, end_date: str = None):
        """
        Initialize the attribution engine.
        
        Args:
            time_range_days: Number of days to look back for attribution analysis (used if start_date/end_date not provided)
            db_config: PostgreSQL database configuration dictionary
            start_date: Start date for analysis (YYYY-MM-DD format)
            end_date: End date for analysis (YYYY-MM-DD format)
        """
        self.time_range_days = time_range_days
        self.db_config = db_config or {}
        self.start_date = start_date
        self.end_date = end_date
        self.orders_df = None
        self.ads_df = None
        self.attribution_results = []
        self.engine = None
        
        # Channel mapping for UTM sources
        # This creates a comprehensive Shopify attributed dataset with:
        # - Meta: Facebook, Instagram, and other Meta platforms (for joining with Meta ad tables)
        # - Google: Google Ads, Google Search, Google Shopping (for joining with Google ad tables)
        # - Organic: Natural search traffic, SEO, unpaid search (for organic analysis)
        # - Direct: Direct visits, bookmarks, email, referrals (for direct traffic analysis)
        self.channel_mapping = {
            'google': 'Google',
            'an': 'Google',       # Google Analytics
            'Google': 'Google',   # Capitalized Google
            'facebook': 'Meta',   # Facebook
            'Facebook': 'Meta',   # Capitalized Facebook
            'fb': 'Meta',         # Facebook shorthand
            'instagram': 'Meta',  # Instagram
            'ig': 'Meta',         # Instagram shorthand
            'IGShopping': 'Meta', # Instagram Shopping
            'shopify': 'Organic', # Shopify keyword goes to Organic
            'duckduckgo': 'Organic', # DuckDuckGo is organic search
            'organic': 'Organic', # Organic channel
            'search': 'Organic',  # Search as organic
            'direct': 'Direct',   # Direct channel
            '{{site_source_name}}': 'Meta', # Site source goes to Meta
            'null': 'Direct/Unknown',
            '': 'Direct/Unknown'
        }
        
        logging.info(f"Attribution Engine initialized with {time_range_days} days lookback")
    
    def connect_to_database(self) -> None:
        """
        Connect to PostgreSQL database using the provided configuration.
        """
        try:
            # PostgreSQL connection
            connection_string = (
                f"postgresql://{self.db_config['user']}:{self.db_config['password']}@"
                f"{self.db_config['host']}:{self.db_config['port']}/{self.db_config['database']}"
            )
            self.engine = create_engine(connection_string, pool_pre_ping=True)
            
            logging.info("Connected to PostgreSQL database successfully")
            
        except Exception as e:
            logging.error(f"Error connecting to PostgreSQL database: {e}")
            raise
    
    def load_data_from_db(self) -> None:
        """
        Load data from PostgreSQL database tables.
        """
        try:
            if not self.engine:
                self.connect_to_database()
            
            # Determine date range for filtering
            if self.start_date and self.end_date:
                start_date_str = self.start_date
                end_date_str = self.end_date
                logging.info(f"Using date range: {start_date_str} to {end_date_str}")
            else:
                # Calculate cutoff date based on time_range_days
                cutoff_date = datetime.now() - timedelta(days=self.time_range_days)
                start_date_str = cutoff_date.strftime('%Y-%m-%d')
                end_date_str = datetime.now().strftime('%Y-%m-%d')
                logging.info(f"Using {self.time_range_days} days lookback: {start_date_str} to {end_date_str}")
            
            logging.info("Loading Shopify orders data from PostgreSQL...")
            
            # Load orders data with line items and product variants
            orders_query = f"""
                SELECT 
                    o.order_id, o.order_name, o.created_at, o.processed_at, o.updated_at,
                    o.closed_at, o.cancelled_at, o.total_price_amount, o.total_price_currency,
                    o.total_shipping_amount, o.total_shipping_currency, o.display_fulfillment_status,
                    o.ship_city, o.ship_province, o.ship_country, o.customer_referrer_url,
                    o.customer_utm_source, o.customer_utm_medium, o.customer_utm_campaign,
                    o.customer_utm_content, o.customer_utm_term, o.created_at_ist,
                    o.processed_at_ist, o.updated_at_ist, o.closed_at_ist, o.cancelled_at_ist,
                    o.customer_journey, o.custom_attributes,
                    -- Line item details
                    li.item_id, li.title as product_title, li.quantity, 
                    li.original_unit_price_amount, li.original_unit_price_currency,
                    li.discounted_unit_price_amount, li.discounted_unit_price_currency,
                    li.variant_id,
                    -- Product variant details
                    pv.sku, pv.variant_title, pv.unit_cost_amount, pv.unit_cost_currency,
                    pv.product_id, pv.product_title as variant_product_title, pv.vendor,
                    pv.selling_price
                FROM shopify_orders o
                LEFT JOIN shopify_order_line_items li ON o.order_id = li.order_id
                LEFT JOIN shopify_product_variants pv ON li.variant_id = pv.variant_id
                WHERE o.created_at_ist::timestamp >= '{start_date_str} 00:00:00+05:30'::timestamp 
                  AND o.created_at_ist::timestamp <= '{end_date_str} 23:59:59+05:30'::timestamp
                  AND o.cancelled_at_ist IS NULL
                ORDER BY o.created_at_ist::timestamp DESC
            """
            
            self.orders_df = pd.read_sql(orders_query, self.engine)
            logging.info(f"Loaded {len(self.orders_df)} orders from PostgreSQL")
            
            # Debug: Check if orders data is empty
            if len(self.orders_df) == 0:
                logging.warning("No orders data loaded! Check date range and database connection.")
            else:
                logging.info(f"Orders date range (IST): {self.orders_df['created_at_ist'].min()} to {self.orders_df['created_at_ist'].max()}")
                logging.info(f"Sample order columns: {list(self.orders_df.columns)}")
            
            logging.info("Loading ad insights data from PostgreSQL...")
            
            # Load ads data with optimized query
            ads_query = f"""
                SELECT 
                    campaign_id, campaign_name, adset_id, adset_name, ad_id, ad_name,
                    date_start, date_stop, hourly_window,
                    impressions, clicks, spend, cpm, cpc, ctr,
                    action_onsite_web_purchase, value_onsite_web_purchase,
                    action_onsite_web_add_to_cart, action_onsite_web_initiate_checkout,
                    action_onsite_web_view_content, action_link_click
                FROM ads_insights_hourly 
                WHERE date_start >= '{start_date_str}'::date AND date_start <= '{end_date_str}'::date
                ORDER BY date_start DESC
            """
            
            self.ads_df = pd.read_sql(ads_query, self.engine)
            logging.info(f"Loaded {len(self.ads_df)} ad records from PostgreSQL")
            
            # Debug: Check if ads data is empty
            if len(self.ads_df) == 0:
                logging.warning("No ads data loaded! Check date range and database connection.")
            else:
                logging.info(f"Ads date range: {self.ads_df['date_start'].min()} to {self.ads_df['date_start'].max()}")
                logging.info(f"Sample ads columns: {list(self.ads_df.columns)}")
            
            # Convert date columns
            self.orders_df['created_at'] = pd.to_datetime(self.orders_df['created_at'])
            self.ads_df['date_start'] = pd.to_datetime(self.ads_df['date_start'])
            
            logging.info(f"Data loaded successfully: {len(self.orders_df)} orders and {len(self.ads_df)} ad records")
            
        except Exception as e:
            logging.error(f"Error loading data from PostgreSQL: {e}")
            raise
    
    def load_data(self) -> None:
        """
        Load and prepare the data from PostgreSQL database.
        """
        # Load from database
        self.load_data_from_db()


        # --- helpers used by create_granular_insights ---
    def _fix_id(self, val):
        """
        Normalize numeric-like IDs to a plain string, avoiding scientific notation.
        Keeps 'Unknown' as-is; returns 'Unknown' for blanks/NaN.
        """
        if val in [None, '', 'Unknown'] or (isinstance(val, float) and np.isnan(val)):
            return 'Unknown'
        try:
            # handles "1.23456e+12" and numeric types
            return str(int(float(val)))
        except (ValueError, TypeError):
            return str(val)

    def _normalize_hourly_window(self, s):
        """
        Convert various hourly window formats to 'HH:00:00 - HH:59:59'.
        Returns None if it can't parse an hour.
        """
        if s is None or (isinstance(s, float) and np.isnan(s)):
            return None
        s = str(s)
        # try to find the first hour number
        m = re.search(r'(\d{1,2})', s)
        if not m:
            return None
        hh = int(m.group(1)) % 24
        return f"{hh:02d}:00:00 - {hh:02d}:59:59"

    def _format_hour_window(self, ts):
        """
        Given a timestamp, return its hour label 'HH:00:00 - HH:59:59'.
        """
        if ts is None:
            return None
        ts = pd.to_datetime(ts, errors='coerce')
        if pd.isna(ts):
            return None
        hh = int(ts.hour)
        return f"{hh:02d}:00:00 - {hh:02d}:59:59"


    def parse_customer_journey(self, journey_str: str) -> Optional[Dict]:
        """
        Parse customer journey JSON string and extract attribution data.
        
        Args:
            journey_str: JSON string containing customer journey data
            
        Returns:
            Dictionary with attribution data or None if parsing fails
        """
        if (journey_str is None or 
            journey_str == '' or 
            str(journey_str).strip() == '' or 
            str(journey_str).strip() == 'null' or
            str(journey_str).strip() == 'None'):
            return None
            
        try:
            # Handle different JSON formats
            if isinstance(journey_str, str):
                journey_data = json.loads(journey_str)
            else:
                journey_data = journey_str
                
            if not journey_data or 'moments' not in journey_data:
                return None
                
            moments = journey_data['moments']
            if not moments:
                return None
                
            # Find the last valid moment with UTM parameters
            for moment in reversed(moments):
                if moment and 'utmParameters' in moment and moment['utmParameters']:
                    utm_params = moment['utmParameters']
                    
                    # Look for content ID first, then campaign, then medium
                    content_id = utm_params.get('content')
                    campaign_id = utm_params.get('campaign')
                    medium_id = utm_params.get('medium')
                    
                    # Return the first valid ID found
                    if content_id and content_id != 'null' and not content_id.startswith('{{'):
                        return {
                            'source': utm_params.get('source'),
                            'medium': utm_params.get('medium'),
                            'campaign': utm_params.get('campaign'),
                            'content': content_id,
                            'term': utm_params.get('term'),
                            'attribution_id': content_id,
                            'attribution_type': 'content'
                        }
                    elif campaign_id and campaign_id != 'null' and not campaign_id.startswith('{{'):
                        return {
                            'source': utm_params.get('source'),
                            'medium': utm_params.get('medium'),
                            'campaign': campaign_id,
                            'content': utm_params.get('content'),
                            'term': utm_params.get('term'),
                            'attribution_id': campaign_id,
                            'attribution_type': 'campaign'
                        }
                    elif medium_id and medium_id != 'null' and not medium_id.startswith('{{'):
                        return {
                            'source': utm_params.get('source'),
                            'medium': medium_id,
                            'campaign': utm_params.get('campaign'),
                            'content': utm_params.get('content'),
                            'term': utm_params.get('term'),
                            'attribution_id': medium_id,
                            'attribution_type': 'medium'
                        }
            
            return None
            
        except (json.JSONDecodeError, KeyError, TypeError) as e:
            logging.warning(f"Error parsing customer journey: {e}")
            return None
    
    def parse_custom_attributes(self, attributes_str: str) -> Optional[Dict]:
        """
        Parse custom attributes JSON string and extract UTM data.
        
        Args:
            attributes_str: JSON string containing custom attributes
            
        Returns:
            Dictionary with UTM data or None if parsing fails
        """
        if (attributes_str is None or 
            attributes_str == '' or 
            str(attributes_str).strip() == '' or 
            str(attributes_str).strip() == 'null' or
            str(attributes_str).strip() == 'None' or
            str(attributes_str).strip() == '[]'):
            return None
            
        try:
            if isinstance(attributes_str, str):
                attributes = json.loads(attributes_str)
            else:
                attributes = attributes_str
                
            if not attributes:
                return None
                
            # Extract UTM parameters from custom attributes
            utm_data = {}
            for attr in attributes:
                if isinstance(attr, dict) and 'key' in attr and 'value' in attr:
                    key = attr['key']
                    value = attr['value']
                    
                    if key.startswith('utm_'):
                        utm_data[key] = value
                    elif key == 'fbclid':
                        utm_data['fbclid'] = value
                    elif key == 'gclid':
                        utm_data['gclid'] = value
            
            if not utm_data:
                return None
                
            # Determine attribution ID (content > campaign > medium)
            attribution_id = None
            attribution_type = None
            
            if 'utm_content' in utm_data and utm_data['utm_content']:
                attribution_id = utm_data['utm_content']
                attribution_type = 'content'
            elif 'utm_campaign' in utm_data and utm_data['utm_campaign']:
                attribution_id = utm_data['utm_campaign']
                attribution_type = 'campaign'
            elif 'utm_medium' in utm_data and utm_data['utm_medium']:
                attribution_id = utm_data['utm_medium']
                attribution_type = 'medium'
            
            if attribution_id:
                return {
                    'source': utm_data.get('utm_source'),
                    'medium': utm_data.get('utm_medium'),
                    'campaign': utm_data.get('utm_campaign'),
                    'content': utm_data.get('utm_content'),
                    'term': utm_data.get('utm_term'),
                    'attribution_id': attribution_id,
                    'attribution_type': attribution_type
                }
            
            return None
            
        except (json.JSONDecodeError, KeyError, TypeError) as e:
            logging.warning(f"Error parsing custom attributes: {e}")
            return None
    
    def get_direct_utm_data(self, row: pd.Series) -> Optional[Dict]:
        """
        Extract UTM data from direct columns in the order data.
        
        Args:
            row: Order row from DataFrame
            
        Returns:
            Dictionary with UTM data or None
        """
        utm_data = {}
        
        # Check direct UTM columns
        utm_source = row.get('customer_utm_source')
        if utm_source is not None and utm_source != '' and str(utm_source).strip() != '':
            utm_data['source'] = utm_source
            
        utm_medium = row.get('customer_utm_medium')
        if utm_medium is not None and utm_medium != '' and str(utm_medium).strip() != '':
            utm_data['medium'] = utm_medium
            
        utm_campaign = row.get('customer_utm_campaign')
        if utm_campaign is not None and utm_campaign != '' and str(utm_campaign).strip() != '':
            utm_data['campaign'] = utm_campaign
            
        utm_content = row.get('customer_utm_content')
        if utm_content is not None and utm_content != '' and str(utm_content).strip() != '':
            utm_data['content'] = utm_content
            
        utm_term = row.get('customer_utm_term')
        if utm_term is not None and utm_term != '' and str(utm_term).strip() != '':
            utm_data['term'] = utm_term
        
        if not utm_data:
            return None
        
        # Determine attribution ID
        attribution_id = None
        attribution_type = None
        
        if 'content' in utm_data and utm_data['content']:
            attribution_id = utm_data['content']
            attribution_type = 'content'
        elif 'campaign' in utm_data and utm_data['campaign']:
            attribution_id = utm_data['campaign']
            attribution_type = 'campaign'
        elif 'medium' in utm_data and utm_data['medium']:
            attribution_id = utm_data['medium']
            attribution_type = 'medium'
        
        if attribution_id:
            return {
                'source': utm_data.get('source'),
                'medium': utm_data.get('medium'),
                'campaign': utm_data.get('campaign'),
                'content': utm_data.get('content'),
                'term': utm_data.get('term'),
                'attribution_id': attribution_id,
                'attribution_type': attribution_type
            }
        
        return None
    
    def rollup_ads_to_daily(self) -> pd.DataFrame:
        """
        Roll up hourly ad data to daily level.
        
        Returns:
            DataFrame with daily aggregated ad data
        """
        logging.info("Rolling up hourly ad data to daily level...")
        
        # Group by date and aggregate metrics
        daily_ads = self.ads_df.groupby([
            'date_start', 'campaign_id', 'campaign_name', 'adset_id', 'adset_name', 'ad_id', 'ad_name'
        ]).agg({
            'impressions': 'sum',
            'clicks': 'sum',
            'spend': 'sum',
            'action_onsite_web_purchase': 'sum',
            'value_onsite_web_purchase': 'sum',
            'action_onsite_web_add_to_cart': 'sum',
            'action_onsite_web_initiate_checkout': 'sum',
            'action_onsite_web_view_content': 'sum',
            'action_link_click': 'sum'
        }).reset_index()
        
        # Calculate daily metrics with zero division protection
        daily_ads['cpm'] = np.where(daily_ads['impressions'] > 0, 
                                   (daily_ads['spend'] / daily_ads['impressions'] * 1000), 0)
        daily_ads['cpc'] = np.where(daily_ads['clicks'] > 0, 
                                   (daily_ads['spend'] / daily_ads['clicks']), 0)
        daily_ads['ctr'] = np.where(daily_ads['impressions'] > 0, 
                                   (daily_ads['clicks'] / daily_ads['impressions'] * 100), 0)
        
        logging.info(f"Rolled up to {len(daily_ads)} daily ad records")
        return daily_ads
    
    def map_attribution_to_campaign(self, attribution_data: Dict, daily_ads: pd.DataFrame, channel: str = None) -> Optional[Dict]:
        """
        Map attribution data to campaign information.
        If no match found but channel is known, return "Unknown Campaign" data
        
        Args:
            attribution_data: Attribution data from order
            daily_ads: Daily aggregated ad data
            channel: Channel name (for creating "Unknown Campaign" entries)
            
        Returns:
            Dictionary with campaign mapping or None
        """
        if attribution_data is None or 'attribution_id' not in attribution_data:
            # If we have a channel but no attribution data, create "Unknown Campaign" entry
            if channel and channel in ['Meta', 'Google', 'Organic']:
                logging.debug(f"Creating 'Unknown Campaign' entry for {channel} with no attribution data")
                return {
                    'campaign_id': None,
                    'campaign_name': f"Unknown Campaign - {channel}",
                    'adset_id': None,
                    'adset_name': f"Unknown Adset - {channel}",
                    'ad_id': None,
                    'ad_name': f"Unknown Ad - {channel}",
                    'impressions': 0,
                    'clicks': 0,
                    'spend': 0,
                    'ad_purchases': 0,
                    'ad_revenue': 0
                }
            return None
        
        attribution_id = attribution_data['attribution_id']
        attribution_type = attribution_data.get('attribution_type', 'content')
        
        # Debug: Log matching attempt
        logging.debug(f"Trying to match attribution_id: {attribution_id} (type: {attribution_type})")
        
        # Try to match by different ID types based on attribution type
        matches = []
        
        # Convert attribution_id to string for comparison
        attribution_id_str = str(attribution_id)
        
        if attribution_type == 'content':
            # Match by ad_id (content) - try both string and integer matching
            ad_matches = daily_ads[
                (daily_ads['ad_id'].astype(str) == attribution_id_str) |
                (daily_ads['ad_id'] == attribution_id)
            ]
            if not ad_matches.empty:
                matches.extend(ad_matches.to_dict('records'))
                logging.debug(f"Found {len(ad_matches)} ad matches for content_id: {attribution_id}")
        
        elif attribution_type == 'campaign':
            # Match by campaign_id - try both string and integer matching
            campaign_matches = daily_ads[
                (daily_ads['campaign_id'].astype(str) == attribution_id_str) |
                (daily_ads['campaign_id'] == attribution_id)
            ]
            if not campaign_matches.empty:
                matches.extend(campaign_matches.to_dict('records'))
                logging.debug(f"Found {len(campaign_matches)} campaign matches for campaign_id: {attribution_id}")
        
        elif attribution_type == 'medium':
            # Match by adset_id (medium level) - try both string and integer matching
            adset_matches = daily_ads[
                (daily_ads['adset_id'].astype(str) == attribution_id_str) |
                (daily_ads['adset_id'] == attribution_id)
            ]
            if not adset_matches.empty:
                matches.extend(adset_matches.to_dict('records'))
                logging.debug(f"Found {len(adset_matches)} adset matches for adset_id: {attribution_id}")
        
        if not matches:
            # If no match found but we have a channel, create "Unknown Campaign" entry
            if channel and channel in ['Meta', 'Google', 'Organic']:
                logging.debug(f"No match found for attribution_id {attribution_id} ({attribution_type}), creating 'Unknown Campaign' for {channel}")
                return {
                    'campaign_id': None,
                    'campaign_name': f"Unknown Campaign - {channel}",
                    'adset_id': None,
                    'adset_name': f"Unknown Adset - {channel}",
                    'ad_id': None,
                    'ad_name': f"Unknown Ad - {channel}",
                    'impressions': 0,
                    'clicks': 0,
                    'spend': 0,
                    'ad_purchases': 0,
                    'ad_revenue': 0
                }
            else:
                logging.debug(f"No matches found for attribution_id: {attribution_id}")
            return None
        
        # Take the first match (most recent or highest spend)
        match = matches[0]
        
        return {
            'campaign_id': match['campaign_id'],
            'campaign_name': match['campaign_name'],
            'adset_id': match['adset_id'],
            'adset_name': match['adset_name'],
            'ad_id': match['ad_id'],
            'ad_name': match['ad_name'],
            'impressions': match['impressions'],
            'clicks': match['clicks'],
            'spend': match['spend'],
            'ad_purchases': match['action_onsite_web_purchase'],
            'ad_revenue': match['value_onsite_web_purchase']
        }
    
    def determine_channel(self, source: str, medium: str = None, campaign: str = None, content: str = None, term: str = None) -> str:
        """
        Enhanced channel classification with better organic traffic detection.
        
        Args:
            source: UTM source value
            medium: UTM medium value
            campaign: UTM campaign value
            content: UTM content value
            term: UTM term value
            
        Returns:
            Channel name
        """
        # Convert all parameters to lowercase strings for comparison
        source_lower = str(source).lower().strip() if source else ''
        medium_lower = str(medium).lower().strip() if medium else ''
        campaign_lower = str(campaign).lower().strip() if campaign else ''
        content_lower = str(content).lower().strip() if content else ''
        term_lower = str(term).lower().strip() if term else ''
        
        # Debug logging for channel classification
        if source_lower or medium_lower:
            logging.debug(f"Channel classification - Source: '{source_lower}', Medium: '{medium_lower}', Campaign: '{campaign_lower}', Content: '{content_lower}', Term: '{term_lower}'")
        
        # PRIORITY 1: Check for campaign data - if there's a campaign, it's likely paid advertising
        if campaign_lower and campaign_lower not in ['', 'null', 'none']:
            # If there's a campaign, determine if it's Google or Meta based on source
            if 'google' in source_lower or 'google' in medium_lower or 'googleadservices' in source_lower:
                final_channel = 'Google'
            elif any(indicator in source_lower for indicator in ['facebook', 'fb', 'instagram', 'ig', 'meta', 'igshopping']):
                final_channel = 'Meta'
            else:
                # Default to Meta for campaigns with unknown source (likely Meta)
                final_channel = 'Meta'
            logging.debug(f"Campaign detected: '{campaign_lower}' → Channel: '{final_channel}'")
            return final_channel
        
        # Check for explicit organic indicators
        organic_indicators = ['organic', 'search', 'natural', 'seo', 'unpaid']
        if any(indicator in medium_lower for indicator in organic_indicators):
            return 'Organic'
        if any(indicator in source_lower for indicator in organic_indicators):
            return 'Organic'
        
        # Check for search engine traffic (likely organic)
        search_engines = ['bing', 'yahoo', 'duckduckgo', 'baidu', 'yandex', 'qwant']
        if any(engine in source_lower for engine in search_engines):
            return 'Organic'
        
        # Check for Google traffic (distinguish between paid and organic)
        if 'google' in source_lower or 'google' in medium_lower:
            # If there's a campaign parameter, it's likely paid Google Ads
            if campaign_lower and campaign_lower not in ['', 'null', 'none']:
                return 'Google'
            # If there's content or term, it's likely paid
            elif content_lower or term_lower:
                return 'Google'
            # Otherwise, likely organic Google search
            else:
                return 'Organic'
        
        # Check for Google Ad Services (definitely paid)
        if 'googleadservices' in source_lower or 'googleadservices' in medium_lower:
            return 'Google'
        
        # Check for Meta platforms (including specific UTM sources)
        meta_indicators = ['facebook', 'fb', 'instagram', 'ig', 'meta', 'igshopping']
        if any(indicator in source_lower for indicator in meta_indicators):
            return 'Meta'
        
        # Special case: {{site_source_name}} placeholder goes to Meta
        if source_lower == '{{site_source_name}}':
            return 'Meta'
        
        # Check for Google platforms (including Analytics)
        google_indicators = ['google', 'an']
        if any(indicator in source_lower for indicator in google_indicators):
            return 'Google'
        
        # Check for direct traffic indicators
        direct_indicators = ['direct', 'none', 'null', '']
        if source_lower in direct_indicators or medium_lower in direct_indicators:
            # If we have some meaningful source but it's marked as direct/none/null, 
            # it's likely organic traffic that wasn't properly tracked
            if source_lower and source_lower not in ['', 'null', 'none'] and len(source_lower) > 1:
                return 'Organic'
            else:
                return 'Direct'
        
        # Check for referral traffic that might be organic
        if medium_lower == 'referral':
            # If source is a known search engine, classify as organic
            if any(engine in source_lower for engine in search_engines):
                return 'Organic'
            # Otherwise, keep as referral (could be organic partnerships)
            return 'Organic'
        
        # If no UTM parameters but has some source, analyze the source
        if source_lower and source_lower not in ['', 'null', 'none']:
            # Check if it looks like a search engine domain
            if any(domain in source_lower for domain in ['.com', '.co.', '.org', '.net', '.in', '.uk']):
                # If it's a search engine, likely organic
                if any(engine in source_lower for engine in search_engines + ['google']):
                    return 'Organic'
                # If it's a social platform, likely Meta
                elif any(platform in source_lower for platform in meta_indicators):
                    return 'Meta'
                # Otherwise, could be organic referral
                else:
                    return 'Organic'
        
        # Additional logic: if we have medium but no source, analyze medium patterns
        if medium_lower and medium_lower not in ['', 'null', 'none']:
            # Check for organic indicators in medium
            if medium_lower in ['search', 'organic', 'natural', 'seo', 'unpaid']:
                return 'Organic'
            # Check for social indicators
            elif medium_lower in ['social', 'social-media', 'socialmedia']:
                return 'Meta'
            # Check for referral
            elif medium_lower == 'referral':
                return 'Organic'  # Referrals are often organic
            # Check for email
            elif medium_lower == 'email':
                return 'Direct'  # Email is direct marketing
            # Check for cpc (cost per click - paid)
            elif medium_lower == 'cpc':
                # Need to check source to determine if Google or Meta
                if 'google' in source_lower:
                    return 'Google'
                elif any(platform in source_lower for platform in meta_indicators):
                    return 'Meta'
                else:
                    return 'Google'  # Default to Google for CPC
        
        # Default fallback
        if not source_lower or source_lower in ['', 'null', 'none']:
            final_channel = 'Direct'
        else:
            # If we have some source but can't classify, it's likely organic traffic
            # This includes orders that might have been misclassified as Direct
            final_channel = 'Organic'
        
        # Log the final channel decision
        if source_lower or medium_lower:
            logging.debug(f"Final channel decision: '{final_channel}' for Source: '{source_lower}', Medium: '{medium_lower}'")
        
        return final_channel
    
    def infer_channel_from_order_patterns(self, order: pd.Series) -> str:
        """
        Infer channel from order patterns when no explicit attribution data is available.
        
        Args:
            order: Order data row
            
        Returns:
            Inferred channel name
        """
        # Check for referrer URL patterns
        referrer_url = order.get('customer_referrer_url', '')
        if referrer_url and str(referrer_url).strip() not in ['', 'null', 'None']:
            referrer_lower = str(referrer_url).lower()
            
            # Check for search engines
            search_engines = ['google.com', 'bing.com', 'yahoo.com', 'duckduckgo.com']
            if any(engine in referrer_lower for engine in search_engines):
                return 'Organic'
            
            # Check for social media platforms
            social_platforms = ['facebook.com', 'instagram.com', 'twitter.com', 'linkedin.com']
            if any(platform in referrer_lower for platform in social_platforms):
                return 'Meta'
            
            # Check for Shopify and other e-commerce platforms (organic referral)
            if 'shopify.com' in referrer_lower:
                return 'Organic'
            
            # Check for other websites (likely organic referral)
            if any(domain in referrer_lower for domain in ['.com', '.co.', '.org', '.net']):
                return 'Organic'
        
        # Check for customer journey patterns
        customer_journey = order.get('customer_journey', '')
        if customer_journey and str(customer_journey).strip() not in ['', 'null', 'None']:
            journey_lower = str(customer_journey).lower()
            
            # Look for organic indicators in journey
            organic_indicators = ['search', 'organic', 'seo', 'unpaid', 'natural', 'shopify']
            if any(indicator in journey_lower for indicator in organic_indicators):
                return 'Organic'
            
            # Look for social indicators
            social_indicators = ['social', 'facebook', 'instagram', 'fb', 'ig', 'igshopping']
            if any(indicator in journey_lower for indicator in social_indicators):
                return 'Meta'
        
        # Check for custom attributes patterns
        custom_attrs = order.get('custom_attributes', '')
        if custom_attrs and str(custom_attrs).strip() not in ['', 'null', 'None', '[]']:
            attrs_lower = str(custom_attrs).lower()
            
            # Look for organic indicators
            organic_indicators = ['search', 'organic', 'seo', 'unpaid', 'natural', 'shopify']
            if any(indicator in attrs_lower for indicator in organic_indicators):
                return 'Organic'
            
            # Look for social indicators
            social_indicators = ['social', 'facebook', 'instagram', 'fb', 'ig', 'igshopping']
            if any(indicator in attrs_lower for indicator in social_indicators):
                return 'Meta'
        
        # If no patterns found, default to Direct
        return 'Direct'
    
    def process_attribution(self) -> pd.DataFrame:
        """
        Process attribution for all orders.
        
        Returns:
            DataFrame with attribution results
        """
        logging.info("Starting attribution processing...")
        
        # Roll up ads to daily level
        daily_ads = self.rollup_ads_to_daily()
        
        # Debug: Log available ad data
        logging.info(f"Available ad data: {len(daily_ads)} records")
        if not daily_ads.empty:
            logging.info(f"Ad ID range: {daily_ads['ad_id'].min()} to {daily_ads['ad_id'].max()}")
            logging.info(f"Campaign ID range: {daily_ads['campaign_id'].min()} to {daily_ads['campaign_id'].max()}")
            logging.info(f"Adset ID range: {daily_ads['adset_id'].min()} to {daily_ads['adset_id'].max()}")
            logging.info(f"Sample ad names: {daily_ads['ad_name'].head().tolist()}")
            logging.info(f"Sample campaign names: {daily_ads['campaign_name'].head().tolist()}")
        
        results = []
        
        # Group orders by order_id to handle multiple line items per order
        orders_grouped = self.orders_df.groupby('order_id')
        total_orders = len(orders_grouped)
        
        logging.info(f"Processing {total_orders} unique orders with line items...")
        
        for order_idx, (order_id, order_group) in enumerate(orders_grouped):
            try:
                # Get the first row for order-level data (all rows have same order data)
                order = order_group.iloc[0]
                order_date = order['created_at_ist']
                order_value = float(order['total_price_amount'])
                
                # Debug: Log first few orders to see data structure
                if order_idx < 3:
                    logging.info(f"Processing order {order_id}: {len(order_group)} line items, customer_journey type={type(order.get('customer_journey'))}, "
                               f"custom_attributes type={type(order.get('custom_attributes'))}")
                
                # Initialize attribution data
                attribution_data = None
                attribution_source = None
                campaign_data = None
                
                # 1. Try customer journey first (highest priority)
                customer_journey = order.get('customer_journey')
                if (customer_journey is not None and 
                    customer_journey != '' and 
                    str(customer_journey).strip() != '' and 
                    str(customer_journey).strip() != 'null' and
                    str(customer_journey).strip() != 'None'):
                    attribution_data = self.parse_customer_journey(customer_journey)
                    if attribution_data:
                        attribution_source = 'customer_journey'
                
                # 2. Try custom attributes if no customer journey
                if attribution_data is None:
                    custom_attributes = order.get('custom_attributes')
                    if (custom_attributes is not None and 
                        custom_attributes != '' and 
                        str(custom_attributes).strip() != '' and 
                        str(custom_attributes).strip() != 'null' and
                        str(custom_attributes).strip() != '[]' and
                        str(custom_attributes).strip() != 'None'):
                        attribution_data = self.parse_custom_attributes(custom_attributes)
                    if attribution_data:
                        attribution_source = 'custom_attributes'
                
                # 3. Try direct UTM columns as fallback
                if attribution_data is None:
                    attribution_data = self.get_direct_utm_data(order)
                    if attribution_data:
                        attribution_source = 'direct_utm'
                
                # Determine channel first (needed for "Unknown Campaign" logic)
                source = attribution_data.get('source') if attribution_data else None
                medium = attribution_data.get('medium') if attribution_data else None
                campaign = attribution_data.get('campaign') if attribution_data else None
                content = attribution_data.get('content') if attribution_data else None
                term = attribution_data.get('term') if attribution_data else None
                channel = self.determine_channel(source, medium, campaign, content, term)
                
                # Log channel classification for debugging
                if order_idx < 5:  # Log first 5 orders for debugging
                    logging.info(f"Order {order_id} channel classification: Source='{source}', Medium='{medium}', Campaign='{campaign}' → Channel='{channel}'")
                
                # Map to campaign data if attribution found
                if attribution_data:
                    # Debug: Log attribution data for first few orders
                    if order_idx < 5:
                        logging.info(f"Order {order_id} attribution: {attribution_data}")
                    campaign_data = self.map_attribution_to_campaign(attribution_data, daily_ads, channel)
                    if campaign_data and order_idx < 5:
                        logging.info(f"Order {order_id} campaign match: {campaign_data.get('campaign_name', 'N/A')} - {campaign_data.get('ad_name', 'N/A')}")
                else:
                    # If no attribution data but we have a channel, create "Unknown Campaign" entry
                    if channel and channel in ['Meta', 'Google', 'Organic', 'Direct']:
                        # Special case: If channel is Meta but no attribution, classify as Organic
                        if channel == 'Meta':
                            channel = 'Organic'
                            if order_idx < 5:  # Log first 5 orders for debugging
                                logging.info(f"Order {order_id} Meta channel reclassified as Organic (no attribution)")
                        campaign_data = self.map_attribution_to_campaign(None, daily_ads, channel)
                    else:
                        # Try to infer channel from order patterns if no explicit attribution
                        inferred_channel = self.infer_channel_from_order_patterns(order)
                        if inferred_channel:
                            channel = inferred_channel
                            if order_idx < 5:  # Log first 5 orders for debugging
                                logging.info(f"Order {order_id} channel inferred from patterns: '{inferred_channel}'")
                        campaign_data = self.map_attribution_to_campaign(None, daily_ads, channel)
                
                # Calculate order-level financial metrics and collect SKU information
                total_cogs = 0
                line_items_info = []
                skus_in_order = []
                sku_quantities = {}
                
                for _, line_item in order_group.iterrows():
                    # Calculate COGS for this line item
                    quantity = line_item.get('quantity', 0) or 0
                    unit_cost = line_item.get('unit_cost_amount', 0) or 0
                    line_item_cogs = quantity * unit_cost
                    total_cogs += line_item_cogs
                    
                    # Calculate revenue for this line item
                    unit_price = line_item.get('discounted_unit_price_amount', 0) or line_item.get('original_unit_price_amount', 0) or 0
                    line_item_revenue = quantity * unit_price
                    
                    # Collect SKU information
                    sku = line_item.get('sku')
                    if sku:
                        skus_in_order.append(sku)
                        sku_quantities[sku] = sku_quantities.get(sku, 0) + quantity
                    
                    # Store line item info
                    line_items_info.append({
                        'sku': sku,
                        'product_title': line_item.get('product_title'),
                        'quantity': quantity,
                        'unit_price': unit_price,
                        'unit_cost': unit_cost,
                        'line_item_revenue': line_item_revenue,
                        'line_item_cogs': line_item_cogs
                    })
                
                # Create result record with order-level data and financial metrics
                result = {
                    # Order Information (essential for analysis)
                    'order_id': order_id,
                    'order_name': order['order_name'],
                    'order_date': order_date,
                    'order_value': order_value,
                    'order_currency': order.get('total_price_currency'),
                    'ship_city': order.get('ship_city'),
                    'ship_province': order.get('ship_province'),
                    'ship_country': order.get('ship_country'),
                    
                    # Financial Metrics
                    'total_cogs': total_cogs,
                    'line_items_count': len(order_group),
                    
                    # SKU Information
                    'skus': ', '.join(set(skus_in_order)) if skus_in_order else None,
                    'unique_skus_count': len(set(skus_in_order)) if skus_in_order else 0,
                    'total_sku_quantity': sum(sku_quantities.values()) if sku_quantities else 0,
                    
                    # Attribution Information (derived/calculated)
                    'channel': channel,
                    'attribution_source': attribution_source,
                    'attribution_id': attribution_data.get('attribution_id') if attribution_data else None,
                    'attribution_type': attribution_data.get('attribution_type') if attribution_data else None,
                    
                    # UTM Parameters (parsed from JSON)
                    'utm_source': attribution_data.get('source') if attribution_data else None,
                    'utm_medium': attribution_data.get('medium') if attribution_data else None,
                    'utm_campaign': attribution_data.get('campaign') if attribution_data else None,
                    'utm_content': attribution_data.get('content') if attribution_data else None,
                    'utm_term': attribution_data.get('term') if attribution_data else None,
                    
                    # Campaign Data (essential for analysis)
                    'campaign_id': campaign_data.get('campaign_id') if campaign_data else None,
                    'campaign_name': campaign_data.get('campaign_name') if campaign_data else None,
                    'adset_id': campaign_data.get('adset_id') if campaign_data else None,
                    'adset_name': campaign_data.get('adset_name') if campaign_data else None,
                    'ad_id': campaign_data.get('ad_id') if campaign_data else None,
                    'ad_name': campaign_data.get('ad_name') if campaign_data else None,
                    
                    # Investigation flags (derived)
                    'has_customer_journey': pd.notna(order.get('customer_journey')),
                    'has_custom_attributes': pd.notna(order.get('custom_attributes'))
                }
                
                results.append(result)
                
                # Log progress
                if (order_idx + 1) % 100 == 0:
                    logging.info(f"Processed {order_idx + 1}/{total_orders} orders...")
                
            except Exception as e:
                logging.error(f"Error processing order {order_id}: {e}")
                continue
        
        logging.info(f"Completed attribution processing for {len(results)} orders")
        
        # Debug: Count successful matches
        if results:
            results_df_temp = pd.DataFrame(results)
            matched_orders = results_df_temp[results_df_temp['campaign_name'].notna()].shape[0]
            total_orders = len(results_df_temp)
            logging.info(f"Attribution matching summary: {matched_orders}/{total_orders} orders matched to campaigns ({matched_orders/total_orders*100:.1f}%)")
        
        # Debug: Check if we have any results
        if len(results) == 0:
            logging.warning("No attribution results generated! Check if orders data is loaded properly.")
            # Create empty DataFrame with expected columns
            results_df = pd.DataFrame(columns=[
                'order_id', 'order_name', 'order_date', 'order_value', 'order_currency',
                'ship_city', 'ship_province', 'ship_country', 'total_cogs', 'line_items_count', 
                'skus', 'unique_skus_count', 'total_sku_quantity', 'channel', 'attribution_source', 
                'attribution_id', 'attribution_type', 'utm_source', 'utm_medium', 'utm_campaign', 
                'utm_content', 'utm_term', 'campaign_id', 'campaign_name', 'adset_id', 'adset_name', 
                'ad_id', 'ad_name', 'has_customer_journey', 'has_custom_attributes', 'is_attributed', 
                'is_paid_channel'
            ])
        else:
            # Create DataFrame and add summary statistics
            results_df = pd.DataFrame(results)
        
        # Add summary columns
        results_df['is_attributed'] = results_df['attribution_source'].notna().astype(bool)
        results_df['is_paid_channel'] = results_df['channel'].isin(['Meta', 'Google']).astype(bool)
        
        # POST-PROCESSING: Reclassify Direct orders with campaign data to Meta
        # This ensures consistency between process_attribution and granular insights
        # Ensure proper data types for logical operations
        channel_mask = results_df['channel'] == 'Direct'
        utm_campaign_mask = (results_df['utm_campaign'].notna() & 
                           (results_df['utm_campaign'].astype(str) != '') & 
                           (results_df['utm_campaign'].astype(str) != 'null'))
        attribution_source_mask = (results_df['attribution_source'].notna() & 
                                 (results_df['attribution_source'].astype(str) != ''))
        campaign_name_mask = (results_df['campaign_name'].notna() & 
                            (results_df['campaign_name'].astype(str) != '') & 
                            ~results_df['campaign_name'].astype(str).str.contains('Unknown', na=False))
        
        direct_with_campaign_mask = (channel_mask & 
                                   (utm_campaign_mask | attribution_source_mask | campaign_name_mask))
        if direct_with_campaign_mask.any():
            logging.info(f"Reclassifying {direct_with_campaign_mask.sum()} Direct orders with campaign/attribution data to Meta in process_attribution")
            results_df.loc[direct_with_campaign_mask, 'channel'] = 'Meta'
            # Update is_paid_channel flag
            results_df.loc[direct_with_campaign_mask, 'is_paid_channel'] = True
        
        # POST-PROCESSING: Reclassify Direct orders with source data to Organic
        # This catches orders that should be Organic but were classified as Direct
        # Ensure proper data types for logical operations
        channel_mask = results_df['channel'] == 'Direct'
        utm_source_mask = (results_df['utm_source'].notna() & 
                          (results_df['utm_source'].astype(str) != '') & 
                          (results_df['utm_source'].astype(str) != 'null'))
        utm_medium_mask = (results_df['utm_medium'].notna() & 
                          (results_df['utm_medium'].astype(str) != '') & 
                          (results_df['utm_medium'].astype(str) != 'null'))
        
        direct_with_source_mask = (channel_mask & (utm_source_mask | utm_medium_mask))
        
        if direct_with_source_mask.any():
            logging.info(f"Reclassifying {direct_with_source_mask.sum()} Direct orders with source data to Organic in process_attribution")
            results_df.loc[direct_with_source_mask, 'channel'] = 'Organic'
        
        # Log channel distribution summary
        if not results_df.empty and 'channel' in results_df.columns:
            channel_summary = results_df['channel'].value_counts()
            logging.info("Channel classification summary:")
            for channel, count in channel_summary.items():
                logging.info(f"  {channel}: {count} orders")
            
            # Log some examples of each channel
            for channel in ['Meta', 'Google', 'Organic', 'Direct']:
                if channel in results_df['channel'].values:
                    sample_orders = results_df[results_df['channel'] == channel].head(3)
                    logging.info(f"Sample {channel} orders:")
                    for _, order in sample_orders.iterrows():
                        logging.info(f"  Order {order['order_id']}: UTM Source='{order.get('utm_source', 'N/A')}', UTM Medium='{order.get('utm_medium', 'N/A')}'")
            
            # Debug: Check if we have organic and Google orders
            organic_orders = results_df[results_df['channel'] == 'Organic']
            google_orders = results_df[results_df['channel'] == 'Google']
            meta_orders = results_df[results_df['channel'] == 'Meta']
            direct_orders = results_df[results_df['channel'] == 'Direct']
            logging.info(f"Organic orders count: {len(organic_orders)}")
            logging.info(f"Google orders count: {len(google_orders)}")
            logging.info(f"Meta orders count: {len(meta_orders)}")
            logging.info(f"Direct orders count: {len(direct_orders)}")
            if len(organic_orders) > 0:
                logging.info(f"Sample organic order UTM: Source='{organic_orders.iloc[0].get('utm_source', 'N/A')}', Medium='{organic_orders.iloc[0].get('utm_medium', 'N/A')}'")
            if len(google_orders) > 0:
                logging.info(f"Sample Google order UTM: Source='{google_orders.iloc[0].get('utm_source', 'N/A')}', Medium='{google_orders.iloc[0].get('utm_medium', 'N/A')}'")
            if len(meta_orders) > 0:
                logging.info(f"Sample Meta order UTM: Source='{meta_orders.iloc[0].get('utm_source', 'N/A')}', Medium='{meta_orders.iloc[0].get('utm_medium', 'N/A')}', Campaign='{meta_orders.iloc[0].get('utm_campaign', 'N/A')}'")
            if len(direct_orders) > 0:
                logging.info(f"Sample Direct order UTM: Source='{direct_orders.iloc[0].get('utm_source', 'N/A')}', Medium='{direct_orders.iloc[0].get('utm_medium', 'N/A')}', Campaign='{direct_orders.iloc[0].get('utm_campaign', 'N/A')}'")
        
        # Convert timezone-aware datetimes to timezone-naive for Excel compatibility
        if 'order_date' in results_df.columns:
            results_df['order_date'] = pd.to_datetime(results_df['order_date']).dt.tz_localize(None)
        
        return results_df
    
    def create_investigation_data(self) -> pd.DataFrame:
        """
        Create investigation data with raw JSON samples for debugging.
        
        Returns:
            DataFrame with raw data samples
        """
        logging.info("Creating investigation data with raw JSON samples...")
        
        investigation_data = []
        
        # Sample orders for investigation (mix of attributed and unattributed)
        if len(self.orders_df) == 0:
            logging.info("No orders data available for investigation")
            return pd.DataFrame()
            
        sample_size = min(50, len(self.orders_df))  # Max 50 samples
        if sample_size == 0:
            return pd.DataFrame()
            
        step = max(1, len(self.orders_df) // sample_size) if sample_size > 0 else 1
        sample_indices = list(range(0, len(self.orders_df), step))[:sample_size]
        
        for idx in sample_indices:
            order = self.orders_df.iloc[idx]
            
            # Get raw JSON data
            customer_journey_raw = order.get('customer_journey', '')
            custom_attributes_raw = order.get('custom_attributes', '')
            
            # Try to parse and format JSON for readability
            try:
                if (customer_journey_raw is not None and 
                    customer_journey_raw != '' and 
                    str(customer_journey_raw).strip() != '' and
                    str(customer_journey_raw).strip() != 'null' and
                    str(customer_journey_raw).strip() != 'None'):
                    customer_journey_formatted = json.dumps(json.loads(customer_journey_raw), indent=2)
                else:
                    customer_journey_formatted = "No data"
            except:
                customer_journey_formatted = str(customer_journey_raw)[:500] + "..." if len(str(customer_journey_raw)) > 500 else str(customer_journey_raw)
            
            try:
                if (custom_attributes_raw is not None and 
                    custom_attributes_raw != '' and 
                    str(custom_attributes_raw).strip() != '' and
                    str(custom_attributes_raw).strip() != 'null' and
                    str(custom_attributes_raw).strip() != 'None' and
                    str(custom_attributes_raw).strip() != '[]'):
                    custom_attributes_formatted = json.dumps(json.loads(custom_attributes_raw), indent=2)
                else:
                    custom_attributes_formatted = "No data"
            except:
                custom_attributes_formatted = str(custom_attributes_raw)[:500] + "..." if len(str(custom_attributes_raw)) > 500 else str(custom_attributes_raw)
            
            investigation_record = {
                'order_id': order['order_id'],
                'order_name': order['order_name'],
                'order_date': pd.to_datetime(order['created_at']).tz_localize(None) if pd.notna(order['created_at']) else None,
                'order_value': order['total_price_amount'],
                'customer_journey_raw': customer_journey_formatted,
                'custom_attributes_raw': custom_attributes_formatted,
                'customer_utm_source': order.get('customer_utm_source'),
                'customer_utm_medium': order.get('customer_utm_medium'),
                'customer_utm_campaign': order.get('customer_utm_campaign'),
                'customer_utm_content': order.get('customer_utm_content'),
                'customer_utm_term': order.get('customer_utm_term'),
                'customer_referrer_url': order.get('customer_referrer_url'),
                'has_customer_journey': (customer_journey_raw is not None and 
                                       customer_journey_raw != '' and 
                                       str(customer_journey_raw).strip() != '' and
                                       str(customer_journey_raw).strip() != 'null' and
                                       str(customer_journey_raw).strip() != 'None'),
                'has_custom_attributes': (custom_attributes_raw is not None and 
                                        custom_attributes_raw != '' and 
                                        str(custom_attributes_raw).strip() != '' and
                                        str(custom_attributes_raw).strip() != 'null' and
                                        str(custom_attributes_raw).strip() != 'None' and
                                        str(custom_attributes_raw).strip() != '[]'),
                'journey_length': len(str(customer_journey_raw)),
                'attributes_length': len(str(custom_attributes_raw))
            }
            
            investigation_data.append(investigation_record)
        
        return pd.DataFrame(investigation_data)
    
    def generate_summary_report(self, results_df: pd.DataFrame) -> Dict:
        """
        Generate summary statistics for the attribution analysis.
        
        Args:
            results_df: DataFrame with attribution results
            
        Returns:
            Dictionary with summary statistics
        """
        logging.info("Generating summary report...")
        
        # Debug: Check data types and column contents
        if not results_df.empty:
            logging.info(f"Results DataFrame shape: {results_df.shape}")
            logging.info(f"Results DataFrame columns: {list(results_df.columns)}")
            if 'is_attributed' in results_df.columns:
                logging.info(f"is_attributed column dtype: {results_df['is_attributed'].dtype}")
                logging.info(f"is_attributed unique values: {results_df['is_attributed'].unique()}")
                logging.info(f"is_attributed sample values: {results_df['is_attributed'].head().tolist()}")
        
        total_orders = len(results_df)
        total_value = results_df['order_value'].sum()
        
        # Attribution rates
        # Ensure is_attributed is boolean type for safe operations
        if not results_df.empty and 'is_attributed' in results_df.columns:
            # Convert to boolean if needed and handle any mixed types
            is_attributed_series = results_df['is_attributed'].astype(bool)
            attributed_orders = is_attributed_series.sum()
        else:
            attributed_orders = 0
        attribution_rate = (attributed_orders / total_orders * 100) if total_orders > 0 else 0
        
        # Channel breakdown
        if not results_df.empty:
            channel_breakdown = results_df.groupby('channel').agg({
                'order_id': 'count',
                'order_value': 'sum'
            }).rename(columns={'order_id': 'orders', 'order_value': 'revenue'})
            channel_breakdown['percentage'] = (channel_breakdown['orders'] / total_orders * 100)
        else:
            channel_breakdown = pd.DataFrame(columns=['orders', 'revenue', 'percentage'])
        
        # Attribution source breakdown
        if not results_df.empty:
            source_breakdown = results_df.groupby('attribution_source').agg({
                'order_id': 'count',
                'order_value': 'sum'
            }).rename(columns={'order_id': 'orders', 'order_value': 'revenue'})
            source_breakdown['percentage'] = (source_breakdown['orders'] / total_orders * 100)
        else:
            source_breakdown = pd.DataFrame(columns=['orders', 'revenue', 'percentage'])
        
        # Campaign performance (for attributed orders)
        if not results_df.empty and 'is_attributed' in results_df.columns:
            # Ensure is_attributed is boolean for safe filtering
            is_attributed_mask = results_df['is_attributed'].astype(bool)
            attributed_df = results_df[is_attributed_mask]
        else:
            attributed_df = pd.DataFrame()
        campaign_performance = None
        if not attributed_df.empty:
            # Get daily ads data for performance metrics
            daily_ads = self.rollup_ads_to_daily()
            
            # Group by campaign level for total sales summary
            campaign_performance = attributed_df.groupby(['campaign_name', 'adset_name', 'ad_name', 'channel']).agg({
                'order_id': 'count',
                'order_value': 'sum',
                'total_cogs': 'sum'
            }).rename(columns={'order_id': 'orders', 'order_value': 'total_sales'})
            
                        # Join with ads data to get ad performance metrics (impressions, clicks, spend only)
            if not daily_ads.empty:
                # Create a key for joining - AGGREGATE ads data first to avoid duplicates
                campaign_performance = campaign_performance.reset_index()
                
                # Debug: Log before aggregation
                logging.info(f"Before aggregation - daily_ads shape: {daily_ads.shape}")
                logging.info(f"Before aggregation - campaign_performance shape: {campaign_performance.shape}")
                
                # Aggregate ads data by campaign, adset, and ad names to avoid duplicates
                # NOTE: We only use ad performance metrics, NOT Facebook's purchase/revenue data
                daily_ads_aggregated = daily_ads.groupby(['campaign_name', 'adset_name', 'ad_name']).agg({
                    'impressions': 'sum',
                    'clicks': 'sum', 
                    'spend': 'sum'
                }).reset_index()
                
                # Debug: Log after aggregation
                logging.info(f"After aggregation - daily_ads_aggregated shape: {daily_ads_aggregated.shape}")
                
                # Join on campaign, adset, and ad names
                campaign_performance = campaign_performance.merge(
                    daily_ads_aggregated, 
                    on=['campaign_name', 'adset_name', 'ad_name'], 
                    how='left'
                )
                
                # Debug: Log after join
                logging.info(f"After join - campaign_performance shape: {campaign_performance.shape}")
                
                # Rename columns to match expected format
                campaign_performance = campaign_performance.rename(columns={
                    'impressions': 'ad_impressions',
                    'clicks': 'ad_clicks', 
                    'spend': 'ad_spend'
                })
                
                # Fill NaN values with 0 for performance metrics
                performance_cols = ['ad_impressions', 'ad_clicks', 'ad_spend']
                campaign_performance[performance_cols] = campaign_performance[performance_cols].fillna(0)
                
                # Add Shopify-based purchase and revenue data (from our attribution results)
                campaign_performance['shopify_purchases'] = campaign_performance['orders']  # Use actual order count
                campaign_performance['shopify_revenue'] = campaign_performance['total_sales']  # Use actual Shopify revenue
            else:
                # If no ads data, add empty performance columns
                campaign_performance = campaign_performance.reset_index()
                campaign_performance['ad_impressions'] = 0
                campaign_performance['ad_clicks'] = 0
                campaign_performance['ad_spend'] = 0
                campaign_performance['ad_purchases'] = 0
                campaign_performance['ad_revenue'] = 0
            
            # Calculate performance metrics with zero division protection
            # Use Shopify data for sales metrics, Facebook data only for ad performance
            campaign_performance['roas'] = np.where(campaign_performance['ad_spend'] > 0, 
                                                   (campaign_performance['shopify_revenue'] / campaign_performance['ad_spend']), 0)
            campaign_performance['ctr'] = np.where(campaign_performance['ad_impressions'] > 0, 
                                                  (campaign_performance['ad_clicks'] / campaign_performance['ad_impressions'] * 100), 0)
            campaign_performance['conversion_rate'] = np.where(campaign_performance['ad_clicks'] > 0, 
                                                              (campaign_performance['shopify_purchases'] / campaign_performance['ad_clicks'] * 100), 0)
            campaign_performance['avg_order_value'] = np.where(campaign_performance['shopify_purchases'] > 0, 
                                                              (campaign_performance['shopify_revenue'] / campaign_performance['shopify_purchases']), 0)
            
            # Sort by total sales descending
            campaign_performance = campaign_performance.sort_values('total_sales', ascending=False)
        
        summary = {
            'total_orders': total_orders,
            'total_revenue': total_value,
            'attributed_orders': attributed_orders,
            'attribution_rate': attribution_rate,
            'channel_breakdown': channel_breakdown,
            'source_breakdown': source_breakdown,
            'campaign_performance': campaign_performance
        }
        
        logging.info(f"Summary: {total_orders} orders, {attribution_rate:.1f}% attribution rate")
        
        return summary
    
    def create_hierarchical_sku_performance(self, ad_sku_performance: pd.DataFrame, ad_performance: pd.DataFrame) -> pd.DataFrame:
        """
        Create hierarchical SKU performance data in the format: Campaign > Adset > Ad > SKU
        with Total rows for aggregation.
        
        Args:
            ad_sku_performance: DataFrame with ad-SKU performance data
            ad_performance: DataFrame with ad performance data
            
        Returns:
            DataFrame in hierarchical format with Total rows
        """
        logging.info("Creating hierarchical SKU performance data...")
        
        if ad_sku_performance.empty:
            return pd.DataFrame()
        
        try:
            # Ensure we have the required columns
            required_cols = ['campaign_name', 'adset_name', 'ad_name', 'sku', 'sku_revenue', 'sku_cogs', 'orders']
            missing_cols = [col for col in required_cols if col not in ad_sku_performance.columns]
            if missing_cols:
                logging.warning(f"Missing columns in ad_sku_performance: {missing_cols}")
                return pd.DataFrame()
            
            # Create a simplified hierarchical structure
            hierarchical_data = []
            
            # Sort by campaign, adset, ad, then by SKU revenue
            ad_sku_sorted = ad_sku_performance.sort_values([
                'campaign_name', 'adset_name', 'ad_name', 'sku_revenue'
            ], ascending=[True, True, True, False])
            
            current_campaign = None
            current_adset = None
            current_ad = None
            
            # Track totals for each level
            campaign_totals = {}
            adset_totals = {}
            ad_totals = {}
            
            for _, row in ad_sku_sorted.iterrows():
                campaign_name = str(row['campaign_name']) if pd.notna(row['campaign_name']) else 'Unknown'
                adset_name = str(row['adset_name']) if pd.notna(row['adset_name']) else 'Unknown'
                ad_name = str(row['ad_name']) if pd.notna(row['ad_name']) else 'Unknown'
                sku = str(row['sku']) if pd.notna(row['sku']) else 'Unknown'
                
                # Initialize totals
                if campaign_name not in campaign_totals:
                    campaign_totals[campaign_name] = {'total_sales': 0, 'total_cogs': 0, 'ad_spent': 0, 'orders': 0}
                if (campaign_name, adset_name) not in adset_totals:
                    adset_totals[(campaign_name, adset_name)] = {'total_sales': 0, 'total_cogs': 0, 'ad_spent': 0, 'orders': 0}
                if (campaign_name, adset_name, ad_name) not in ad_totals:
                    ad_totals[(campaign_name, adset_name, ad_name)] = {'total_sales': 0, 'total_cogs': 0, 'ad_spent': 0, 'orders': 0}
                
                # Add blank row when campaign changes
                if current_campaign is not None and current_campaign != campaign_name:
                    hierarchical_data.append(self.create_blank_row())
                
                # Add blank row when adset changes
                if (current_campaign == campaign_name and 
                    current_adset is not None and current_adset != adset_name):
                    hierarchical_data.append(self.create_blank_row())
                
                # Add blank row when ad changes
                if (current_campaign == campaign_name and 
                    current_adset == adset_name and 
                    current_ad is not None and current_ad != ad_name):
                    hierarchical_data.append(self.create_blank_row())
                
                # Get ad performance data for this ad
                ad_perf = ad_performance[
                    (ad_performance['campaign_name'] == campaign_name) &
                    (ad_performance['adset_name'] == adset_name) &
                    (ad_performance['ad_name'] == ad_name)
                ]
                
                ad_spend = ad_perf['ad_spend'].iloc[0] if not ad_perf.empty else 0
                ad_clicks = ad_perf['ad_clicks'].iloc[0] if not ad_perf.empty else 0
                ad_ctr = ad_perf['ctr'].iloc[0] if not ad_perf.empty else 0
                
                # Add SKU row
                sku_row = {
                    'sku': sku,
                    'campaign_name': campaign_name if current_campaign != campaign_name else '',
                    'adset_name': adset_name if (current_campaign != campaign_name or current_adset != adset_name) else '',
                    'ad_name': ad_name if (current_campaign != campaign_name or current_adset != adset_name or current_ad != ad_name) else '',
                    'ctr': round(float(ad_ctr), 2) if ad_ctr else 0,
                    'purchase_roas': round(float(row.get('sku_roas', 0)), 2) if pd.notna(row.get('sku_roas')) else 0,
                    'Clicks': int(ad_clicks) if ad_clicks else 0,
                    'Add to Cart': 0,  # Placeholder - would need to be calculated from order data
                    'Conversion Rate (%)': round((float(row['orders']) / max(float(ad_clicks), 1)) * 100, 2) if ad_clicks else 0,
                    'total_sales': int(float(row['sku_revenue'])) if pd.notna(row['sku_revenue']) else 0,
                    'total_cogs': int(float(row['sku_cogs'])) if pd.notna(row['sku_cogs']) else 0,
                    'ad_spent': '',  # Empty for individual SKU rows
                    'net_profit': '',  # Empty for individual SKU rows
                    'orders': int(float(row['orders'])) if pd.notna(row['orders']) else 0
                }
                hierarchical_data.append(sku_row)
                
                # Update totals
                sku_revenue = float(row['sku_revenue']) if pd.notna(row['sku_revenue']) else 0
                sku_cogs = float(row['sku_cogs']) if pd.notna(row['sku_cogs']) else 0
                orders = int(float(row['orders'])) if pd.notna(row['orders']) else 0
                
                ad_totals[(campaign_name, adset_name, ad_name)]['total_sales'] += sku_revenue
                ad_totals[(campaign_name, adset_name, ad_name)]['total_cogs'] += sku_cogs
                ad_totals[(campaign_name, adset_name, ad_name)]['ad_spent'] = ad_spend
                ad_totals[(campaign_name, adset_name, ad_name)]['orders'] += orders
                
                adset_totals[(campaign_name, adset_name)]['total_sales'] += sku_revenue
                adset_totals[(campaign_name, adset_name)]['total_cogs'] += sku_cogs
                adset_totals[(campaign_name, adset_name)]['ad_spent'] += ad_spend
                adset_totals[(campaign_name, adset_name)]['orders'] += orders
                
                campaign_totals[campaign_name]['total_sales'] += sku_revenue
                campaign_totals[campaign_name]['total_cogs'] += sku_cogs
                campaign_totals[campaign_name]['ad_spent'] += ad_spend
                campaign_totals[campaign_name]['orders'] += orders
                
                current_campaign = campaign_name
                current_adset = adset_name
                current_ad = ad_name
            
            # Create final hierarchical structure with totals in proper positions
            final_hierarchical_data = []
            current_campaign = None
            current_adset = None
            current_ad = None
            
            for _, row in ad_sku_sorted.iterrows():
                campaign_name = str(row['campaign_name']) if pd.notna(row['campaign_name']) else 'Unknown'
                adset_name = str(row['adset_name']) if pd.notna(row['adset_name']) else 'Unknown'
                ad_name = str(row['ad_name']) if pd.notna(row['ad_name']) else 'Unknown'
                sku = str(row['sku']) if pd.notna(row['sku']) else 'Unknown'
                
                # Add blank row when campaign changes
                if current_campaign is not None and current_campaign != campaign_name:
                    final_hierarchical_data.append(self.create_blank_row())
                
                # Add blank row when adset changes
                if (current_campaign == campaign_name and 
                    current_adset is not None and current_adset != adset_name):
                    final_hierarchical_data.append(self.create_blank_row())
                
                # Add blank row when ad changes
                if (current_campaign == campaign_name and 
                    current_adset == adset_name and 
                    current_ad is not None and current_ad != ad_name):
                    final_hierarchical_data.append(self.create_blank_row())
                
                # Get ad performance data for this ad
                ad_perf = ad_performance[
                    (ad_performance['campaign_name'] == campaign_name) &
                    (ad_performance['adset_name'] == adset_name) &
                    (ad_performance['ad_name'] == ad_name)
                ]
                
                ad_spend = ad_perf['ad_spend'].iloc[0] if not ad_perf.empty else 0
                ad_clicks = ad_perf['ad_clicks'].iloc[0] if not ad_perf.empty else 0
                ad_ctr = ad_perf['ctr'].iloc[0] if not ad_perf.empty else 0
                
                # Add SKU row
                sku_row = {
                    'sku': sku,
                    'campaign_name': campaign_name if current_campaign != campaign_name else '',
                    'adset_name': adset_name if (current_campaign != campaign_name or current_adset != adset_name) else '',
                    'ad_name': ad_name if (current_campaign != campaign_name or current_adset != adset_name or current_ad != ad_name) else '',
                    'ctr': round(float(ad_ctr), 2) if ad_ctr else 0,
                    'purchase_roas': round(float(row.get('sku_roas', 0)), 2) if pd.notna(row.get('sku_roas')) else 0,
                    'Clicks': int(ad_clicks) if ad_clicks else 0,
                    'Add to Cart': 0,  # Placeholder - would need to be calculated from order data
                    'Conversion Rate (%)': round((float(row['orders']) / max(float(ad_clicks), 1)) * 100, 2) if ad_clicks else 0,
                    'total_sales': int(float(row['sku_revenue'])) if pd.notna(row['sku_revenue']) else 0,
                    'total_cogs': int(float(row['sku_cogs'])) if pd.notna(row['sku_cogs']) else 0,
                    'ad_spent': '',  # Empty for individual SKU rows
                    'net_profit': '',  # Empty for individual SKU rows
                    'orders': int(float(row['orders'])) if pd.notna(row['orders']) else 0
                }
                final_hierarchical_data.append(sku_row)
                
                # Check if this is the last SKU for this ad, adset, or campaign
                # Get next row to check if we need to add totals
                current_idx = ad_sku_sorted.index.get_loc(row.name)
                next_idx = current_idx + 1
                
                # Check if we need to add ad total (last SKU for this ad)
                if (next_idx >= len(ad_sku_sorted) or 
                    ad_sku_sorted.iloc[next_idx]['ad_name'] != ad_name or
                    ad_sku_sorted.iloc[next_idx]['adset_name'] != adset_name or
                    ad_sku_sorted.iloc[next_idx]['campaign_name'] != campaign_name):
                    
                    # Add ad total row
                    ad_key = (campaign_name, adset_name, ad_name)
                    if ad_key in ad_totals:
                        total = ad_totals[ad_key]
                        total_row = {
                            'sku': '',
                            'campaign_name': '',
                            'adset_name': '',
                            'ad_name': '',
                            'ctr': '',
                            'purchase_roas': '',
                            'Clicks': '',
                            'Add to Cart': '',
                            'Conversion Rate (%)': '',
                            'total_sales': int(total['total_sales']),
                            'total_cogs': int(total['total_cogs']),
                            'ad_spent': round(total['ad_spent'], 2),
                            'net_profit': round(total['total_sales'] - total['total_cogs'] - total['ad_spent'], 2),
                            'orders': int(total['orders'])
                        }
                        final_hierarchical_data.append(total_row)
                
                # Check if we need to add adset total (last SKU for this adset)
                if (next_idx >= len(ad_sku_sorted) or 
                    ad_sku_sorted.iloc[next_idx]['adset_name'] != adset_name or
                    ad_sku_sorted.iloc[next_idx]['campaign_name'] != campaign_name):
                    
                    # Add adset total row
                    adset_key = (campaign_name, adset_name)
                    if adset_key in adset_totals:
                        total = adset_totals[adset_key]
                        total_row = {
                            'sku': '',
                            'campaign_name': '',
                            'adset_name': '',
                            'ad_name': '',
                            'ctr': '',
                            'purchase_roas': '',
                            'Clicks': '',
                            'Add to Cart': '',
                            'Conversion Rate (%)': '',
                            'total_sales': int(total['total_sales']),
                            'total_cogs': int(total['total_cogs']),
                            'ad_spent': round(total['ad_spent'], 2),
                            'net_profit': round(total['total_sales'] - total['total_cogs'] - total['ad_spent'], 2),
                            'orders': int(total['orders'])
                        }
                        final_hierarchical_data.append(total_row)
                
                # Check if we need to add campaign total (last SKU for this campaign)
                if (next_idx >= len(ad_sku_sorted) or 
                    ad_sku_sorted.iloc[next_idx]['campaign_name'] != campaign_name):
                    
                    # Add campaign total row
                    if campaign_name in campaign_totals:
                        total = campaign_totals[campaign_name]
                        total_row = {
                            'sku': '',
                            'campaign_name': '',
                            'adset_name': '',
                            'ad_name': '',
                            'ctr': '',
                            'purchase_roas': '',
                            'Clicks': '',
                            'Add to Cart': '',
                            'Conversion Rate (%)': '',
                            'total_sales': int(total['total_sales']),
                            'total_cogs': int(total['total_cogs']),
                            'ad_spent': round(total['ad_spent'], 2),
                            'net_profit': round(total['total_sales'] - total['total_cogs'] - total['ad_spent'], 2),
                            'orders': int(total['orders'])
                        }
                        final_hierarchical_data.append(total_row)
                
                current_campaign = campaign_name
                current_adset = adset_name
                current_ad = ad_name
            
            # Add final blank row
            final_hierarchical_data.append(self.create_blank_row())
            
            return pd.DataFrame(final_hierarchical_data)
            
        except Exception as e:
            logging.error(f"Error creating hierarchical SKU performance: {e}")
            return pd.DataFrame()
    
    def create_blank_row(self) -> Dict:
        """Create a blank row for spacing in hierarchical format."""
        return {
            'sku': '',
            'campaign_name': '',
            'adset_name': '',
            'ad_name': '',
            'ctr': '',
            'purchase_roas': '',
            'Clicks': '',
            'Add to Cart': '',
            'Conversion Rate (%)': '',
            'total_sales': '',
            'total_cogs': '',
            'ad_spent': '',
            'net_profit': '',
            'orders': ''
        }
    
    def merge_hierarchical_cells(self, writer, sheet_name: str, hierarchical_data: pd.DataFrame) -> None:
        """
        Merge cells in the hierarchical structure where the same data appears across multiple rows.
        
        Args:
            writer: Excel writer object
            sheet_name: Name of the sheet to merge cells in
            hierarchical_data: DataFrame with hierarchical data
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Border, Side
            
            # Load the workbook
            workbook = writer.book
            worksheet = workbook[sheet_name]
            
            # Define column mappings (0-indexed)
            col_mapping = {
                'sku': 0,
                'campaign_name': 1,
                'adset_name': 2,
                'ad_name': 3,
                'ctr': 4,
                'purchase_roas': 5,
                'Clicks': 6,
                'Add to Cart': 7,
                'Conversion Rate (%)': 8,
                'total_sales': 9,
                'total_cogs': 10,
                'ad_spent': 11,
                'net_profit': 12,
                'orders': 13
            }
            
            # Track ranges for merging
            merge_ranges = []
            
            # Process each column for merging
            for col_name, col_idx in col_mapping.items():
                if col_name in ['campaign_name', 'adset_name', 'ad_name']:
                    current_value = None
                    start_row = None
                    
                    for row_idx, row in hierarchical_data.iterrows():
                        cell_value = row[col_name]
                        
                        # Skip blank rows and total rows
                        if cell_value == '' or pd.isna(cell_value):
                            continue
                        
                        # If this is a new value, end previous merge and start new one
                        if current_value != cell_value:
                            # End previous merge if exists
                            if current_value is not None and start_row is not None:
                                end_row = row_idx + 1  # +1 because Excel is 1-indexed
                                if end_row > start_row + 1:  # Only merge if more than 1 row
                                    merge_ranges.append((col_idx + 1, start_row + 2, col_idx + 1, end_row + 1))
                            
                            # Start new merge
                            current_value = cell_value
                            start_row = row_idx
                        else:
                            # Continue current merge
                            continue
                    
                    # End final merge if exists
                    if current_value is not None and start_row is not None:
                        end_row = len(hierarchical_data)
                        if end_row > start_row + 1:
                            merge_ranges.append((col_idx + 1, start_row + 2, col_idx + 1, end_row + 1))
            
            # Apply merges
            for merge_range in merge_ranges:
                try:
                    worksheet.merge_cells(start_row=merge_range[1], start_column=merge_range[0],
                                        end_row=merge_range[3], end_column=merge_range[2])
                except Exception as e:
                    logging.warning(f"Could not merge cells {merge_range}: {e}")
            
            # Apply formatting to merged cells
            self.format_hierarchical_sheet(worksheet, hierarchical_data)
            
            logging.info(f"Applied {len(merge_ranges)} cell merges to {sheet_name}")
            
        except Exception as e:
            logging.error(f"Error merging hierarchical cells: {e}")
    
    def format_hierarchical_sheet(self, worksheet, hierarchical_data: pd.DataFrame) -> None:
        """
        Apply formatting to the hierarchical sheet.
        
        Args:
            worksheet: OpenPyXL worksheet object
            hierarchical_data: DataFrame with hierarchical data
        """
        try:
            from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
            
            # Define styles
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            header_font = Font(bold=True)
            total_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
            
            # Apply formatting to all cells
            for row_idx in range(1, len(hierarchical_data) + 2):  # +2 for header
                for col_idx in range(1, 15):  # 14 columns
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    
                    # Apply border
                    cell.border = thin_border
                    
                    # Header row formatting
                    if row_idx == 1:
                        cell.font = header_font
                        cell.alignment = center_alignment
                        cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                    else:
                        # Data row formatting
                        data_row_idx = row_idx - 2  # -2 because of header and 0-indexing
                        if 0 <= data_row_idx < len(hierarchical_data):
                            row_data = hierarchical_data.iloc[data_row_idx]
                            
                            # Check if this is a total row (empty sku, campaign_name, adset_name, ad_name)
                            is_total_row = (row_data['sku'] == '' and 
                                          row_data['campaign_name'] == '' and 
                                          row_data['adset_name'] == '' and 
                                          row_data['ad_name'] == '')
                            
                            if is_total_row:
                                # Total row formatting
                                cell.fill = total_fill
                                cell.font = Font(bold=True)
                                cell.alignment = center_alignment
                            else:
                                # Regular data row formatting
                                if col_idx in [1, 2, 3, 4]:  # Text columns
                                    cell.alignment = left_alignment
                                else:  # Numeric columns
                                    cell.alignment = center_alignment
                                    
                                # Format numeric columns
                                if col_idx in [5, 6, 9, 10, 11, 12, 13]:  # Numeric columns
                                    try:
                                        if cell.value != '' and cell.value is not None:
                                            if col_idx in [5, 6]:  # Decimal columns
                                                cell.number_format = '0.00'
                                            else:  # Integer columns
                                                cell.number_format = '0'
                                    except:
                                        pass
            
            # Auto-adjust column widths
            for col_idx in range(1, 15):
                column_letter = worksheet.cell(row=1, column=col_idx).column_letter
                max_length = 0
                
                for row_idx in range(1, len(hierarchical_data) + 2):
                    cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                
                # Set column width (with some padding)
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logging.info("Applied formatting to hierarchical sheet")
            
        except Exception as e:
            logging.error(f"Error formatting hierarchical sheet: {e}")
    
    def create_detailed_sku_attribution(self, results_df: pd.DataFrame) -> pd.DataFrame:
        """
        Create detailed SKU-level attribution data showing which campaign/adset/ad sold which SKU.
        
        Args:
            results_df: DataFrame with attribution results
            
        Returns:
            DataFrame with detailed SKU attribution data
        """
        logging.info("Creating detailed SKU attribution data...")
        
        if results_df.empty:
            return pd.DataFrame()
        
        detailed_data = []
        
        for _, row in results_df.iterrows():
            if row['skus'] and row['is_attributed']:
                # Get the line items data for this order to calculate proper SKU values
                order_line_items = self.orders_df[self.orders_df['order_id'] == row['order_id']]
                
                # Create a mapping of SKU to its actual revenue, COGS, and quantity
                sku_data_map = {}
                
                for _, line_item in order_line_items.iterrows():
                    sku = line_item.get('sku')
                    if sku:
                        quantity = line_item.get('quantity', 0) or 0
                        unit_price = line_item.get('discounted_unit_price_amount', 0) or line_item.get('original_unit_price_amount', 0) or 0
                        unit_cost = line_item.get('unit_cost_amount', 0) or 0
                        
                        sku_revenue = quantity * unit_price
                        sku_cogs = quantity * unit_cost
                        sku_profit = sku_revenue - sku_cogs
                        
                        sku_data_map[sku] = {
                            'quantity': quantity,
                            'unit_price': unit_price,
                            'unit_cost': unit_cost,
                            'sku_revenue': sku_revenue,
                            'sku_cogs': sku_cogs,
                            'sku_profit': sku_profit,
                            'product_title': line_item.get('product_title', ''),
                            'variant_title': line_item.get('variant_title', '')
                        }
                
                # Get ad spend data for this ad
                ad_spend = 0
                if row['ad_id'] and row['campaign_name'] and row['adset_name'] and row['ad_name']:
                    # Find matching ad in ads data
                    matching_ads = self.ads_df[
                        (self.ads_df['campaign_name'] == row['campaign_name']) &
                        (self.ads_df['adset_name'] == row['adset_name']) &
                        (self.ads_df['ad_name'] == row['ad_name'])
                    ]
                    if not matching_ads.empty:
                        ad_spend = matching_ads['spend'].sum()
                
                # Create detailed record for each SKU in this order
                sku_list = [sku.strip() for sku in row['skus'].split(',')]
                for sku in sku_list:
                    if sku in sku_data_map:
                        sku_data = sku_data_map[sku]
                        
                        detailed_record = {
                            # Order Information
                            'order_id': row['order_id'],
                            'order_name': row['order_name'],
                            'order_date': row['order_date'],
                            'order_value': row['order_value'],
                            
                            # SKU Information
                            'sku': sku,
                            'product_title': sku_data['product_title'],
                            'variant_title': sku_data['variant_title'],
                            'quantity': sku_data['quantity'],
                            'unit_price': sku_data['unit_price'],
                            'unit_cost': sku_data['unit_cost'],
                            'sku_revenue': sku_data['sku_revenue'],
                            'sku_cogs': sku_data['sku_cogs'],
                            'sku_profit': sku_data['sku_profit'],
                            'sku_profit_margin': (sku_data['sku_profit'] / sku_data['sku_revenue'] * 100) if sku_data['sku_revenue'] > 0 else 0,
                            
                            # Campaign Information
                            'campaign_id': row['campaign_id'],
                            'campaign_name': row['campaign_name'],
                            'adset_id': row['adset_id'],
                            'adset_name': row['adset_name'],
                            'ad_id': row['ad_id'],
                            'ad_name': row['ad_name'],
                            
                            # Ad Spend Information
                            'ad_spend': ad_spend,
                            'sku_roas': (sku_data['sku_revenue'] / ad_spend) if ad_spend > 0 else 0,
                            'sku_net_profit': sku_data['sku_profit'] - ad_spend,
                            
                            # Channel and Attribution
                            'channel': row['channel'],
                            'attribution_source': row['attribution_source'],
                            'utm_source': row['utm_source'],
                            'utm_medium': row['utm_medium'],
                            'utm_campaign': row['utm_campaign'],
                            'utm_content': row['utm_content'],
                            'utm_term': row['utm_term']
                        }
                        
                        detailed_data.append(detailed_record)
        
        return pd.DataFrame(detailed_data)
    
    # Ad level summary function completely removed

    def save_results(self, results_df: pd.DataFrame, summary: Dict, output_prefix: str = "attribution_results") -> None:
        """
        Save results to Excel file with two sheets:
        1. Raw_Attribution_Data: Unique orders with SKU column, price, and COGS
        2. Granular_Insights: Shopify order values appended to ad insights raw data format
        
        Args:
            results_df: DataFrame with attribution results
            summary: Summary statistics
            output_prefix: Prefix for output filename
        """
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = f"{output_prefix}_{timestamp}.xlsx"
        
        # Prepare data outside try block to avoid reference errors
        # Sheet 1: Raw Attribution Data (unique orders with SKU column, price, COGS)
        raw_data = results_df.copy()
        
        # Rename columns for clarity in the raw data sheet
        raw_data = raw_data.rename(columns={
            'order_value': 'price',
            'total_cogs': 'cogs',
            'skus': 'sku'  # This already contains comma-separated SKUs
        })
        
        # Select relevant columns for the raw data sheet
        raw_columns = [
            'order_id', 'order_name', 'order_date', 'price', 'cogs', 'sku',
            'campaign_name', 'adset_name', 'ad_name', 'channel', 'attribution_source'
        ]
        
        # Filter to only include columns that exist
        available_columns = [col for col in raw_columns if col in raw_data.columns]
        raw_data_sheet = raw_data[available_columns]
        
        # Ad level sheet removed - no longer needed
        
        # Sheet 2: Granular Insights - Compact Format (Meta, Google, Organic, Direct channels)
        granular_data = self.create_granular_insights(results_df)
        granular_sheet = None
        
        if not granular_data.empty:
            # Debug: Log available columns in granular data
            logging.info(f"Granular data columns: {list(granular_data.columns)}")
            logging.info(f"Granular data shape: {granular_data.shape}")
            
            # Select relevant columns for the granular insights sheet (compact format for joining with ad insight table)
            # Updated to match the actual columns available in the compact granular data
            granular_columns = [
                'campaign_name', 'adset_name', 'ad_name', 'date_start', 'hour_label', 'channel',
                'spend', 'cpm', 'cpc', 'ctr', 'action_onsite_web_purchase', 'value_onsite_web_purchase',
                'shopify_orders', 'shopify_revenue', 'shopify_cogs', 'total_sku_quantity',
                'utm_source', 'utm_medium', 'utm_campaign', 'utm_content', 'utm_term',
                'order_ids', 'attribution_source'
            ]
            
            # Filter to only include columns that exist
            available_granular_columns = [col for col in granular_columns if col in granular_data.columns]
            missing_columns = [col for col in granular_columns if col not in granular_data.columns]
            if missing_columns:
                logging.warning(f"Missing columns in granular data: {missing_columns}")
            
            logging.info(f"Available granular columns: {available_granular_columns}")
            granular_sheet = granular_data[available_granular_columns]
        
        # Create Excel writer
        try:
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                
                # Save raw attribution data
                raw_data_sheet.to_excel(writer, sheet_name='Raw_Attribution_Data', index=False)
                logging.info(f"Saved raw attribution data to sheet 'Raw_Attribution_Data'")
                
                # Ad level sheet removed - no longer needed
                
                # Save granular insights data if available
                if granular_sheet is not None and not granular_sheet.empty:
                    granular_sheet.to_excel(writer, sheet_name='Granular_Insights_Compact', index=False)
                    logging.info(f"Saved compact granular insights to sheet 'Granular_Insights_Compact'")
                else:
                    logging.info("No granular insights data to save")
            
            logging.info(f"Saved attribution results to Excel file: {excel_file}")
            
        except Exception as e:
            logging.error(f"Error saving Excel file: {e}")
            # Fallback to CSV files
            results_file = f"{output_prefix}_{timestamp}.csv"
            raw_data_sheet.to_csv(results_file, index=False)
            logging.info(f"Fallback: Saved raw data to CSV file: {results_file}")
            
            # Ad level CSV removed - no longer needed
            
            if granular_sheet is not None and not granular_sheet.empty:
                granular_file = f"{output_prefix}_granular_{timestamp}.csv"
                granular_sheet.to_csv(granular_file, index=False)
                logging.info(f"Fallback: Saved granular insights to CSV file: {granular_file}")
            raise
        
        # Also save individual CSV files for compatibility
        raw_csv_file = f"{output_prefix}_raw_{timestamp}.csv"
        raw_data_sheet.to_csv(raw_csv_file, index=False)
        logging.info(f"Also saved raw data to CSV: {raw_csv_file}")
        
        # Ad level CSV removed - no longer needed
        
        if granular_sheet is not None and not granular_sheet.empty:
            granular_csv_file = f"{output_prefix}_granular_{timestamp}.csv"
            granular_sheet.to_csv(granular_csv_file, index=False)
            logging.info(f"Also saved granular insights to CSV: {granular_csv_file}")
    
    def run_attribution_analysis(self, output_prefix: str = "attribution_results") -> Tuple[pd.DataFrame, Dict]:
        """
        Run complete attribution analysis.
        
        Args:
            output_prefix: Prefix for output filenames
            
        Returns:
            Tuple of (results DataFrame, summary dictionary)
        """
        logging.info("Starting complete attribution analysis...")
        
        # Load data
        self.load_data()
        
        # Process attribution
        results_df = self.process_attribution()
        
        # Generate summary
        summary = self.generate_summary_report(results_df)
        
        # Save results
        self.save_results(results_df, summary, output_prefix)
        
        logging.info("Attribution analysis completed successfully!")
        
        return results_df, summary
    
    def create_granular_insights(self, results_df: pd.DataFrame) -> pd.DataFrame:
        """
        Granular insights at hourly precision WITHOUT duplicates:
        1) Convert each order timestamp to 'HH:00:00 - HH:59:59'
        2) Build exact join keys on both sides
        3) Aggregate orders once per key, then left-join to ads_insights_hourly
        4) Append 'Unknown Campaign - <channel>' rows for any unmatched order keys
        5) Include ALL orders (Meta, Google, Organic, Direct) for comprehensive coverage
        
        Channel Coverage:
        - Meta: Facebook, Instagram, and other Meta platforms
        - Google: Google Ads, Google Search, Google Shopping
        - Organic: Natural search traffic, SEO, unpaid search
        - Direct: Direct visits, bookmarks, email, referrals
        
        Metrics: spend, orders, revenue, COGS (impressions/clicks removed for clean joining)
        """
        logging.info("Creating granular insights (deduped, hour-bucketed)...")

        if results_df.empty or self.ads_df.empty:
            logging.warning("No data for granular insights.")
            return pd.DataFrame()

        # ---------- Prepare ADS (hourly) ----------
        ads = self.ads_df.copy()

        # Normalize IDs to strings
        for col in ['campaign_id', 'adset_id', 'ad_id']:
            if col in ads.columns:
                ads[col] = ads[col].apply(self._fix_id)

        ads['date_only'] = pd.to_datetime(ads['date_start']).dt.date
        # Normalize hourly_window into canonical 'HH:00:00 - HH:59:59'
        ads['hour_label'] = ads['hourly_window'].apply(self._normalize_hourly_window)

        # Keep only true hourly rows (hour_label present)
        hourly_ads = ads[ads['hour_label'].notna()].copy()

        # ---------- Prepare ORDERS ----------
        orders = results_df.copy()

        # Parse order timestamp robustly: examples like "01/07/2025 00:59:44" (dd/mm/yyyy)
        # Your results_df['order_date'] was set from created_at_ist earlier; normalize anyway:
        orders['order_ts'] = pd.to_datetime(orders.get('order_date'), errors='coerce', dayfirst=True)

        # Fallback: pull from original self.orders_df if missing
        missing_ts = orders['order_ts'].isna()
        if missing_ts.any():
            base = self.orders_df[['order_id', 'created_at_ist']].drop_duplicates()
            base['created_at_ist'] = pd.to_datetime(base['created_at_ist'], errors='coerce')
            orders = orders.merge(base, on='order_id', how='left')
            orders.loc[orders['order_ts'].isna(), 'order_ts'] = orders.loc[orders['order_ts'].isna(), 'created_at_ist']

        # Build order date + hour label
        orders['order_date_only'] = orders['order_ts'].dt.date
        orders['hour_label'] = orders['order_ts'].apply(self._format_hour_window)

        # Normalize IDs to strings (important to avoid join explosion)
        for col in ['campaign_id', 'adset_id', 'ad_id']:
            if col in orders.columns:
                orders[col] = orders[col].apply(self._fix_id)

        # Unknown name fallbacks (remain None for the actual key, names used for display)
        orders['campaign_name'] = orders['campaign_name'].fillna('')
        orders['adset_name']    = orders['adset_name'].fillna('')
        orders['ad_name']       = orders['ad_name'].fillna('')

        # Include ALL orders (attributed and non-attributed) for comprehensive channel coverage
        # This ensures organic, Google, and other channels are properly represented
        all_orders = orders.copy()

        # ---------- Aggregate orders ONCE per exact key ----------
        order_key_cols = [
            'campaign_id', 'campaign_name',
            'adset_id', 'adset_name',
            'ad_id', 'ad_name',
            'order_date_only', 'hour_label'
        ]

        def _first_nonnull(series):
            return next((x for x in series if pd.notna(x) and x != ''), None)

        orders_by_key = (
                all_orders
            .groupby(order_key_cols, dropna=False)
            .agg(
                shopify_orders=('order_id', 'nunique'),
                shopify_revenue=('order_value', 'sum'),
                shopify_cogs=('total_cogs', 'sum'),
                total_sku_quantity=('total_sku_quantity', 'sum'),
                channel=('channel', _first_nonnull),
                attribution_source=('attribution_source', _first_nonnull),
                utm_source=('utm_source', _first_nonnull),
                utm_medium=('utm_medium', _first_nonnull),
                utm_campaign=('utm_campaign', _first_nonnull),
                utm_content=('utm_content', _first_nonnull),
                utm_term=('utm_term', _first_nonnull),
                order_ids=('order_id', lambda s: ', '.join(sorted(map(lambda x: str(x), set(s)))))
            )
            .reset_index()
        )

        # ---------- Exact join: ADS x ORDERS ----------
        ads_key_cols = [
            'campaign_id', 'campaign_name',
            'adset_id', 'adset_name',
            'ad_id', 'ad_name',
            'date_only', 'hour_label'
        ]

        orders_by_key = orders_by_key.rename(columns={'order_date_only': 'date_only'})
        hourly_ads_subset = hourly_ads[
            ['campaign_id','campaign_name','adset_id','adset_name','ad_id','ad_name',
                 'date_only','hour_label','spend','cpm','cpc','ctr',
             'action_onsite_web_purchase','value_onsite_web_purchase']
        ].copy()

        # Use a full outer join to ensure ALL orders are included
        # This will include orders that match ads AND orders that don't have ads data
        granular = hourly_ads_subset.merge(
            orders_by_key,
            on=['campaign_id','campaign_name','adset_id','adset_name','ad_id','ad_name','date_only','hour_label'],
            how='outer',
            indicator=True
        )
        
        # Since we used outer join, we can identify unmatched orders using the merge indicator
        # 'right_only' means orders without ads data (organic, Google, etc.)
        # 'left_only' means ads without orders
        # 'both' means orders with ads data

        # Fill NA (ads rows with no matching orders)
        for col, default in [
        ('shopify_orders', 0),
        ('shopify_revenue', 0.0),
        ('shopify_cogs', 0.0),
        ('total_sku_quantity', 0),
        ('channel', ''),                # was None
        ('attribution_source', ''),     # was None
        ('utm_source', ''),             # was None
        ('utm_medium', ''),             # was None
        ('utm_campaign', ''),           # was None
        ('utm_content', ''),            # was None
        ('utm_term', ''),               # was None
        ('order_ids', '')
                    ]:
            if col not in granular.columns:
                granular[col] = default
            else:
                granular[col] = granular[col].fillna(default)


        # Only calculate essential metrics for compact format
        # Removed: shopify_profit, shopify_roas, shopify_conversion_rate, shopify_avg_order_value
        # Removed: Facebook metrics and difference calculations
        
        # Enhance channel detection for better organic/Google identification
        def enhance_channel_detection(row):
            """Enhanced channel detection using the same logic as determine_channel"""
            utm_source = row.get('utm_source', '')
            utm_medium = row.get('utm_medium', '')
            utm_campaign = row.get('utm_campaign', '')
            utm_content = row.get('utm_content', '')
            utm_term = row.get('utm_term', '')
            
            # Use the enhanced determine_channel method
            enhanced_channel = self.determine_channel(utm_source, utm_medium, utm_campaign, utm_content, utm_term)
            
            # Only override if we got a better classification
            if enhanced_channel != 'Direct/Unknown' and enhanced_channel != 'Direct':
                return enhanced_channel
            
            return row.get('channel', '')
        
        # Apply enhanced channel detection
        granular['channel'] = granular.apply(enhance_channel_detection, axis=1)
        
        # POST-PROCESSING: Reclassify Direct orders with campaign data to Meta
        # This catches any orders that slipped through the initial classification
        # Add safety checks for column access
        if all(col in granular.columns for col in ['channel', 'utm_campaign', 'attribution_source', 'campaign_name']):
            # Ensure proper data types for logical operations
            channel_mask = granular['channel'] == 'Direct'
            utm_campaign_mask = (granular['utm_campaign'].notna() & 
                               (granular['utm_campaign'].astype(str) != '') & 
                               (granular['utm_campaign'].astype(str) != 'null'))
            attribution_source_mask = (granular['attribution_source'].notna() & 
                                     (granular['attribution_source'].astype(str) != ''))
            campaign_name_mask = (granular['campaign_name'].notna() & 
                                (granular['campaign_name'].astype(str) != '') & 
                                ~granular['campaign_name'].astype(str).str.contains('Unknown', na=False))
            
            direct_with_campaign_mask = (channel_mask & 
                                       (utm_campaign_mask | attribution_source_mask | campaign_name_mask))
        else:
            direct_with_campaign_mask = pd.Series([False] * len(granular), index=granular.index)
        if direct_with_campaign_mask.any():
            logging.info(f"Reclassifying {direct_with_campaign_mask.sum()} Direct orders with campaign/attribution data to Meta")
            granular.loc[direct_with_campaign_mask, 'channel'] = 'Meta'
            # Update campaign names to reflect Meta classification
            granular.loc[direct_with_campaign_mask, 'campaign_name'] = granular.loc[direct_with_campaign_mask, 'utm_campaign'].apply(
                lambda x: f"Meta Campaign - {x}" if pd.notna(x) and x != '' and x != 'null' else "Meta Campaign - Unknown"
            )
    
        # POST-PROCESSING: Reclassify Direct orders with source data to Organic
        # This catches orders that should be Organic but were classified as Direct
        if all(col in granular.columns for col in ['channel', 'utm_source', 'utm_medium']):
            # Ensure proper data types for logical operations
            channel_mask = granular['channel'] == 'Direct'
            utm_source_mask = (granular['utm_source'].notna() & 
                              (granular['utm_source'].astype(str) != '') & 
                              (granular['utm_source'].astype(str) != 'null'))
            utm_medium_mask = (granular['utm_medium'].notna() & 
                              (granular['utm_medium'].astype(str) != '') & 
                              (granular['utm_medium'].astype(str) != 'null'))
            
            direct_with_source_mask = (channel_mask & (utm_source_mask | utm_medium_mask))
        else:
            direct_with_source_mask = pd.Series([False] * len(granular), index=granular.index)
        
        if direct_with_source_mask.any():
            logging.info(f"Reclassifying {direct_with_source_mask.sum()} Direct orders with source data to Organic")
            granular.loc[direct_with_source_mask, 'channel'] = 'Organic'
            # Update campaign names to reflect Organic classification
            granular.loc[direct_with_source_mask, 'utm_source'].apply(
                lambda x: f"Organic Source - {x}" if pd.notna(x) and x != '' and x != 'null' else "Organic Source - Unknown"
            )

        # ---------- Process unmatched orders (organic, Google, etc.) ----------
        # Since we used outer join, orders without ads data are already in the granular data
        # We just need to clean them up and ensure they have proper campaign names
        
        # For orders without ads data (right_only), update campaign names to show they're unknown
        if '_merge' in granular.columns:
            unmatched_mask = granular['_merge'] == 'right_only'
            if unmatched_mask.any():
                granular.loc[unmatched_mask, 'campaign_name'] = granular.loc[unmatched_mask, 'channel'].apply(
                    lambda ch: f"Unknown Campaign - {ch}" if ch and ch != '' else "Unknown Campaign - Direct/Unknown"
                )
                granular.loc[unmatched_mask, 'adset_name'] = granular.loc[unmatched_mask, 'channel'].apply(
                    lambda ch: f"Unknown Adset - {ch}" if ch and ch != '' else "Unknown Adset - Direct/Unknown"
                )
                granular.loc[unmatched_mask, 'ad_name'] = granular.loc[unmatched_mask, 'channel'].apply(
                    lambda ch: f"Unknown Ad - {ch}" if ch and ch != '' else "Unknown Ad - Direct/Unknown"
                )
                
                # Set ad metrics to 0 for orders without ads data
                ad_metric_cols = ['spend', 'cpm', 'cpc', 'ctr', 'action_onsite_web_purchase', 'value_onsite_web_purchase']
                for col in ad_metric_cols:
                    if col in granular.columns:
                        granular.loc[unmatched_mask, col] = 0.0
        else:
            unmatched_mask = pd.Series([False] * len(granular), index=granular.index)


        # Remove the merge indicator column
        granular = granular.drop(columns=['_merge'])

        # ---------- Final ordering & cleanliness ----------
        # Reconstruct a 'date_start' column for downstream sheets using date_only
        granular['date_start'] = pd.to_datetime(granular['date_only'])
        # Provide expected 'hourly_window' alias for downstream export (sheet expects this name)
        if 'hour_label' in granular.columns and 'hourly_window' not in granular.columns:
            granular['hourly_window'] = granular['hour_label']

        # De-dupe guard: if any duplicates slipped through, collapse Shopify metrics per exact key
        key_cols = ['campaign_id','campaign_name','adset_id','adset_name','ad_id','ad_name','date_start','hour_label']
        numeric_to_sum = [
                'spend',
                'shopify_orders','shopify_revenue','shopify_cogs','total_sku_quantity'
            ]
            # All metrics are now simple counts/values; summing is acceptable if duplicates exist—this block runs rarely.
        if granular.duplicated(key_cols).any():
            logging.info("Duplicate hourly keys detected; collapsing to unique keys.")
            # For strings like order_ids/channel, keep first; for numeric, sum
            keep_first = {
                'channel':'first','attribution_source':'first','utm_source':'first','utm_medium':'first',
                'utm_campaign':'first','utm_content':'first','utm_term':'first','order_ids':'first','date_only':'first'
            }
            agg_map = {c: 'sum' for c in numeric_to_sum if c in granular.columns}
            agg_map.update(keep_first)
            granular = granular.groupby(key_cols, dropna=False, as_index=False).agg(agg_map)

        # Sort by date then hour descending revenue (optional)
        granular = granular.sort_values(['date_start','hour_label','shopify_revenue'], ascending=[True, True, False]).reset_index(drop=True)

        # Log channel distribution for analysis
        try:
            # Ensure numeric columns are properly typed before aggregation
            numeric_cols = ['shopify_orders', 'shopify_revenue', 'shopify_cogs']
            for col in numeric_cols:
                if col in granular.columns:
                    # Convert to numeric, coercing errors to NaN
                    granular[col] = pd.to_numeric(granular[col], errors='coerce').fillna(0)
            
            channel_summary = granular.groupby('channel').agg({
                'shopify_orders': 'sum',
                'shopify_revenue': 'sum',
                'shopify_cogs': 'sum'
            }).round(2)
            
            logging.info(f"Channel distribution in granular insights:")
            for channel, data in channel_summary.iterrows():
                logging.info(f"  {channel}: {data['shopify_orders']} orders, ₹{data['shopify_revenue']:,.2f} revenue")
        except Exception as e:
            logging.error(f"Error in channel distribution analysis: {e}")
            logging.info("Skipping channel distribution analysis due to error")
        
        # Log unmatched orders by channel to ensure organic/Google are included
        try:
            if 'campaign_name' in granular.columns:
                unmatched_mask = granular['campaign_name'].str.contains('Unknown Campaign', na=False)
                if unmatched_mask.any():
                    unmatched_channel_summary = granular[unmatched_mask].groupby('channel').agg({
                        'shopify_orders': 'sum',
                        'shopify_revenue': 'sum'
                    }).round(2)
                    logging.info(f"Unmatched orders by channel (added as Unknown Campaign):")
                    for channel, data in unmatched_channel_summary.iterrows():
                        logging.info(f"  {channel}: {data['shopify_orders']} orders, ₹{data['shopify_revenue']:,.2f} revenue")
        except Exception as e:
            logging.error(f"Error in unmatched orders analysis: {e}")
            logging.info("Skipping unmatched orders analysis due to error")
        
        # Debug: Log some sample orders by channel to see what's happening
        logging.info("Sample orders by channel in granular insights:")
        for channel in ['Meta', 'Google', 'Organic', 'Direct']:
            if channel in granular['channel'].values:
                sample_orders = granular[granular['channel'] == channel].head(2)
                logging.info(f"  {channel} channel sample:")
                for _, order in sample_orders.iterrows():
                    logging.info(f"    Campaign: {order.get('campaign_name', 'N/A')}, Orders: {order.get('shopify_orders', 0)}, Revenue: ₹{order.get('shopify_revenue', 0):,.2f}")
        
        # Log summary of channel reclassification
        meta_orders = granular[granular['channel'] == 'Meta']
        if len(meta_orders) > 0:
            campaign_meta_orders = meta_orders[meta_orders['campaign_name'].str.contains('Meta Campaign', na=False)]
            if len(campaign_meta_orders) > 0:
                logging.info(f"Meta orders with campaign data: {len(campaign_meta_orders)} (reclassified from Direct)")
                logging.info(f"Total Meta orders: {len(meta_orders)}")
        
        # Debug: Check data types of key columns in granular insights:
        logging.info("Data types of key columns in granular insights:")
        for col in ['channel', 'shopify_orders', 'shopify_revenue', 'shopify_cogs']:
            if col in granular.columns:
                logging.info(f"  {col}: {granular[col].dtype}")
                if granular[col].dtype == 'object':
                    logging.info(f"    Sample values: {granular[col].head(3).tolist()}")

        logging.info(f"Granular insights rows: {len(granular)}; Shopify revenue total: {granular['shopify_revenue'].sum():,.2f}")
        return granular



def main():
    """
    Main function to run the attribution analysis.
    """
    print("="*60)
    print("ATTRIBUTION ANALYSIS ENGINE")
    print("="*60)
    print("This script runs attribution analysis using PostgreSQL database tables:")
    print("- shopify_orders")
    print("- ads_insights_hourly")
    print()
    print("NEW FEATURE: Granular Insights (Hourly Precision) - Compact Format")
    print("- Appends Shopify order purchase values to ad insights raw data format")
    print("- Clean, compact format optimized for joining with ad insight tables")
    print("- Comprehensive channel coverage: Meta, Google, Organic, Direct")
    print("- Matches orders to specific hours based on created_at_ist timestamp")
    print("- Essential metrics only: spend, orders, revenue, COGS")
    print()
    
    # Try to load configuration from config.py
    try:
        from config import POSTGRESQL_CONFIG, TIME_RANGE_DAYS, OUTPUT_PREFIX, USE_DATE_RANGE, START_DATE, END_DATE
        DB_CONFIG = POSTGRESQL_CONFIG
        print("✅ Loaded configuration from config.py")
        
        # Display date range configuration
        if USE_DATE_RANGE:
            print(f"📅 Using date range: {START_DATE} to {END_DATE}")
        else:
            print(f"📅 Using {TIME_RANGE_DAYS} days lookback")
            
    except ImportError:
        print("⚠️  config.py not found, using default configuration")
        print("💡 Create config.py with your PostgreSQL credentials for easier setup")
        
        # Default configuration
        TIME_RANGE_DAYS = 30
        OUTPUT_PREFIX = "attribution_analysis"
        USE_DATE_RANGE = False
        START_DATE = None
        END_DATE = None
        DB_CONFIG = {
            'host': 'selericdb.postgres.database.azure.com',
            'port': 5432,
            'database': 'postgres',
            'user': 'admin_seleric',
            'password': 'SelericDB246'
        }
    
    # Ask user for input
    print("Choose an option:")
    print("1. Run attribution analysis")
    print("2. Test database connection only")
    
    choice = input("\nEnter your choice (1 or 2): ").strip()
    
    try:
        if choice == "1":
            print("\n🔍 Testing database connection...")
            if test_database_connection(DB_CONFIG, START_DATE if USE_DATE_RANGE else None, END_DATE if USE_DATE_RANGE else None):
                print("\n🚀 Starting attribution analysis from PostgreSQL...")
                
                # Initialize attribution engine with database config and date range
                if USE_DATE_RANGE:
                    engine = AttributionEngine(
                        time_range_days=TIME_RANGE_DAYS, 
                        db_config=DB_CONFIG,
                        start_date=START_DATE,
                        end_date=END_DATE
                    )
                else:
                    engine = AttributionEngine(
                        time_range_days=TIME_RANGE_DAYS, 
                        db_config=DB_CONFIG
                    )
                
                # Run analysis from database
                results_df, summary = engine.run_attribution_analysis(
                    output_prefix=OUTPUT_PREFIX
                )
                
                # Print summary
                print("\n" + "="*50)
                print("ATTRIBUTION ANALYSIS SUMMARY")
                print("="*50)
                print(f"Total Orders: {summary['total_orders']:,}")
                print(f"Total Revenue: ₹{summary['total_revenue']:,.2f}")
                print(f"Attributed Orders: {summary['attributed_orders']:,}")
                print(f"Attribution Rate: {summary['attribution_rate']:.1f}%")
                
                print("\nChannel Breakdown:")
                print(summary['channel_breakdown'][['orders', 'revenue', 'percentage']].round(2))
                
                print("\nAttribution Source Breakdown:")
                print(summary['source_breakdown'][['orders', 'revenue', 'percentage']].round(2))
                
                print("\n" + "="*50)
                print("✅ Analysis complete! Check the generated Excel file with two sheets:")
                print("   1. Raw_Attribution_Data: Unique orders with SKU information")
                print("   2. Granular_Insights_Compact: Shopify order values with Meta, Google, Organic, Direct channels (hourly precision)")
                print("📄 Log file: attribution_analysis.log")
                
            else:
                print("❌ Database connection failed. Please check your configuration.")
        
        elif choice == "2":
            print("\n🔍 Testing database connection...")
            test_database_connection(DB_CONFIG, START_DATE if USE_DATE_RANGE else None, END_DATE if USE_DATE_RANGE else None)
        
        else:
            print("❌ Invalid choice. Please run the script again and choose 1 or 2.")
        
    except Exception as e:
        logging.error(f"Analysis failed: {e}")
        print(f"❌ Error: {e}")
        print("📄 Check attribution_analysis.log for details")
    
    print("\n" + "="*60)
    print("Thank you for using the Attribution Analysis Engine!")
    print("="*60)




def create_postgresql_config_example():
    """
    Example function showing PostgreSQL configuration.
    """
    # PostgreSQL configuration
    postgres_config = {
        'host': 'localhost',
        'port': 5432,
        'database': 'postgres',
        'user': 'admin_seleric',
        'password': 'SelericDB246'
    }
    
    return postgres_config


def test_database_connection(db_config: Dict, start_date: str = None, end_date: str = None) -> bool:
    """
    Test PostgreSQL connection and verify table structure.
    
    Args:
        db_config: PostgreSQL database configuration dictionary
        start_date: Start date for analysis (YYYY-MM-DD format)
        end_date: End date for analysis (YYYY-MM-DD format)
        
    Returns:
        True if connection successful, False otherwise
    """
    try:
        # Create PostgreSQL connection string
        connection_string = (
            f"postgresql://{db_config['user']}:{db_config['password']}@"
            f"{db_config['host']}:{db_config['port']}/{db_config['database']}"
        )
        
        engine = create_engine(connection_string, pool_pre_ping=True)
        
        # Test connection
        with engine.connect() as conn:
            # Check if tables exist
            tables_query = """
                SELECT table_name 
                FROM information_schema.tables 
                WHERE table_schema = 'public' 
                AND table_name IN ('shopify_orders', 'ads_insights_hourly')
            """
            
            result = conn.execute(text(tables_query))
            tables = [row[0] for row in result]
            
            if 'shopify_orders' not in tables:
                print("❌ Table 'shopify_orders' not found in PostgreSQL database")
                return False
            
            if 'ads_insights_hourly' not in tables:
                print("❌ Table 'ads_insights_hourly' not found in PostgreSQL database")
                return False
            
            # Check table structure
            print("✅ PostgreSQL connection successful!")
            print(f"✅ Found tables: {', '.join(tables)}")
            
            # Get row counts
            orders_count = conn.execute(text("SELECT COUNT(*) FROM shopify_orders")).scalar()
            ads_count = conn.execute(text("SELECT COUNT(*) FROM ads_insights_hourly")).scalar()
            
            print(f"📊 shopify_orders: {orders_count:,} rows")
            print(f"📊 ads_insights_hourly: {ads_count:,} rows")
            
            # Get date ranges
            orders_date_range = conn.execute(text("""
                SELECT MIN(created_at_ist::timestamp), MAX(created_at_ist::timestamp) 
                FROM shopify_orders
            """)).fetchone()
            
            ads_date_range = conn.execute(text("""
                SELECT MIN(date_start), MAX(date_start) 
                FROM ads_insights_hourly
            """)).fetchone()
            
            print(f"📅 shopify_orders date range: {orders_date_range[0]} to {orders_date_range[1]}")
            print(f"📅 ads_insights_hourly date range: {ads_date_range[0]} to {ads_date_range[1]}")
            
            # If date range is specified, show filtered counts
            if start_date and end_date:
                filtered_orders_count = conn.execute(text(f"""
                    SELECT COUNT(*) FROM shopify_orders 
                    WHERE created_at_ist::timestamp >= '{start_date} 00:00:00+05:30'::timestamp 
                      AND created_at_ist::timestamp <= '{end_date} 23:59:59+05:30'::timestamp
                      AND cancelled_at_ist IS NULL
                """)).scalar()
                
                filtered_ads_count = conn.execute(text(f"""
                    SELECT COUNT(*) FROM ads_insights_hourly 
                    WHERE date_start >= '{start_date}'::date AND date_start <= '{end_date}'::date
                """)).scalar()
                
                print(f"📊 Filtered shopify_orders ({start_date} to {end_date}): {filtered_orders_count:,} rows")
                print(f"📊 Filtered ads_insights_hourly ({start_date} to {end_date}): {filtered_ads_count:,} rows")
            
            return True
            
    except Exception as e:
        print(f"❌ PostgreSQL connection failed: {e}")
        return False




if __name__ == "__main__":
    main()


